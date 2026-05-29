"""
═══════════════════════════════════════════════════════════════
  AUDIT MANAGEMENT SYSTEM
  LAN-Based CA FirmHub for Chartered Accountant Firms
═══════════════════════════════════════════════════════════════
"""
import os, uuid, json
from datetime import datetime, timedelta
from pathlib import Path
from flask import Flask, request, jsonify, g, send_file, send_from_directory
from config import HOST, PORT, UPLOAD_DIR, BOOKLET_DIR, EXPORT_DIR, SECRET_KEY, SUBSCRIPTION_TOKEN_HOURS, FIRM_NAME, FIRM_REG_NO, FIRM_SUB_ID, FIRM_EXPIRES
from database import get_db, dict_row, dict_rows, init_db, close_db, init_subscriptions_table, init_v5_tables
from auth import (
    hash_password, verify_password, create_token, login_required,
    require_role, log_action
)
from seed_data import seed
from booklet_generator import generate_booklet

import sys as _sys, os as _os
_static = _os.path.join(
    _os.environ.get('BUNDLE_DIR', _os.path.dirname(_os.path.abspath(__file__))),
    'static'
)
app = Flask(__name__, static_folder=_static)
app.config["SECRET_KEY"] = SECRET_KEY
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50 MB

@app.teardown_appcontext
def teardown(exc):
    close_db(exc)


# ═══════════════════════════════════════════════════════════════
#  TEMP FILE CLEANUP
# ═══════════════════════════════════════════════════════════════
import glob as _glob, os as _os, threading as _thr, time as _time

def _cleanup_temp_xlsx():
    """Delete temp export xlsx files older than 60 seconds."""
    static_dir = _os.path.join(_os.path.dirname(__file__), "static")
    for f in _glob.glob(_os.path.join(static_dir, "tmp*.xlsx")):
        try:
            if _time.time() - _os.path.getmtime(f) > 60:
                _os.unlink(f)
        except Exception:
            pass

@app.after_request
def _after(response):
    """Schedule temp file cleanup after each response."""
    t = _thr.Timer(65, _cleanup_temp_xlsx)
    t.daemon = True
    t.start()
    return response

# ═══════════════════════════════════════════════════════════════
#  SERVE FRONTEND
# ═══════════════════════════════════════════════════════════════
@app.route("/")
def index():
    return send_from_directory("static", "index.html")

@app.route("/static/<path:filename>")
def static_files(filename):
    return send_from_directory("static", filename)


# ═══════════════════════════════════════════════════════════════
#  SUBSCRIPTION GATE
# ═══════════════════════════════════════════════════════════════
@app.route("/api/subscription/verify", methods=["POST"])
def subscription_verify():
    """
    First-run activation: CA firm enters Subscription ID once.
    - Validates against firm_identity.json (must match the baked-in sub_id).
    - Checks expiry date.
    - Marks 'activated=1' so this prompt never appears again.
    - Returns a long-lived subscription token stored locally by the browser.
    """
    import jwt as _jwt
    from datetime import date as _date
    data = request.get_json(silent=True) or {}
    sub_id = str(data.get("subscription_id", "")).strip()

    if not sub_id:
        return jsonify({"detail": "Subscription ID is required"}), 403

    # Must match the baked-in firm identity
    if FIRM_SUB_ID and sub_id != FIRM_SUB_ID:
        print(f"  [SUB] REJECTED — ID mismatch: {sub_id!r} (expected {FIRM_SUB_ID!r})")
        return jsonify({"detail": "Invalid Subscription ID for this installation"}), 403

    db = get_db()
    row = dict_row(db.execute(
        "SELECT * FROM subscription_ids WHERE sub_id=? AND is_active=1", (sub_id,)
    ).fetchone())

    if not row:
        print(f"  [SUB] REJECTED subscription_id: {sub_id!r} — not found or inactive")
        return jsonify({"detail": "Invalid or inactive Subscription ID"}), 403

    # Check expiry
    expires_at = row.get("expires_at") or FIRM_EXPIRES
    if expires_at:
        try:
            exp_date = _date.fromisoformat(expires_at)
            if _date.today() > exp_date:
                print(f"  [SUB] REJECTED — subscription expired on {expires_at}")
                return jsonify({"detail": f"Subscription expired on {expires_at}. Please renew."}), 403
        except ValueError:
            pass

    IST = __import__("datetime").timezone(__import__("datetime").timedelta(hours=5, minutes=30))
    now_ist = __import__("datetime").datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S")

    # Mark as activated (first-run complete) and update usage stats
    db.execute(
        "UPDATE subscription_ids SET last_used_at=?, use_count=use_count+1, activated=1 WHERE id=?",
        (now_ist, row["id"])
    )
    db.commit()

    # Issue a long-lived subscription token
    payload = {
        "sub_verified": True,
        "sub_id": sub_id,
        "firm_name": FIRM_NAME,
        "firm_reg_no": FIRM_REG_NO,
        "expires_at": expires_at or "",
        "exp": datetime.utcnow() + timedelta(hours=SUBSCRIPTION_TOKEN_HOURS),
    }
    token = _jwt.encode(payload, SECRET_KEY, algorithm="HS256")
    print(f"  [SUB] ACTIVATED/VERIFIED subscription_id: {sub_id!r} ({row.get('label','—')})")
    return jsonify({
        "subscription_token": token,
        "valid_hours": SUBSCRIPTION_TOKEN_HOURS,
        "firm_name": FIRM_NAME,
        "firm_reg_no": FIRM_REG_NO,
        "activated": True,
    })


@app.route("/api/firm/identity")
def firm_identity():
    """Return this installation's firm identity (public, no auth needed — baked in at packaging time)."""
    db_exists = __import__("pathlib").Path(__import__("config").DB_PATH).exists()
    activated = False
    if db_exists:
        try:
            db = get_db()
            row = db.execute(
                "SELECT activated FROM subscription_ids WHERE sub_id=? LIMIT 1",
                (FIRM_SUB_ID,)
            ).fetchone()
            if row:
                activated = bool(row[0])
        except Exception:
            pass
    return jsonify({
        "firm_name": FIRM_NAME,
        "firm_reg_no": FIRM_REG_NO,
        "sub_id_set": bool(FIRM_SUB_ID),
        "db_exists": db_exists,
        "activated": activated,
    })
    """Validate an existing subscription token (used on page reload/startup)."""
    import jwt as _jwt
    from datetime import date as _date
    data = request.get_json(silent=True) or {}
    token = str(data.get("subscription_token", "")).strip()
    if not token:
        return jsonify({"valid": False}), 200
    try:
        payload = _jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
        # Also re-check expiry against firm_identity.json (in case admin deactivated)
        expires_at = payload.get("expires_at") or FIRM_EXPIRES
        if expires_at:
            try:
                if _date.today() > _date.fromisoformat(expires_at):
                    return jsonify({"valid": False, "reason": "expired"})
            except ValueError:
                pass
        return jsonify({
            "valid": True,
            "firm_name": FIRM_NAME,
            "firm_reg_no": FIRM_REG_NO,
        })
    except _jwt.PyJWTError:
        return jsonify({"valid": False})


# NOTE: Subscription management is intentionally NOT exposed via the web UI.
# All subscription operations (create, renew, deactivate) must be done
# through admin_generator.py on the admin machine only.



@app.route("/api/auth/login", methods=["POST"])
def login():
    from datetime import date as _date
    data = request.get_json()
    print(f"  [LOGIN] Attempt for username: {data.get('username','(empty)')}")

    # Check subscription validity before allowing any login
    if FIRM_EXPIRES:
        try:
            if _date.today() > _date.fromisoformat(FIRM_EXPIRES):
                print(f"  [LOGIN] BLOCKED — subscription expired {FIRM_EXPIRES}")
                return jsonify({"detail": f"Subscription expired on {FIRM_EXPIRES}. Please contact your administrator to renew."}), 403
        except ValueError:
            pass

    db = get_db()
    user = dict_row(db.execute("SELECT * FROM users WHERE username=?", (data.get("username",""),)).fetchone())
    if not user or not verify_password(data.get("password",""), user["password_hash"]):
        print(f"  [LOGIN] FAILED - invalid credentials")
        return jsonify({"detail": "Invalid credentials"}), 401
    if not user["is_active"]:
        print(f"  [LOGIN] FAILED - account deactivated")
        return jsonify({"detail": "Account deactivated"}), 403

    token = create_token(user["id"], user["role"])
    log_action(db, user["id"], "LOGIN", "User", user["id"], ip=request.remote_addr)
    print(f"  [LOGIN] SUCCESS - {user['full_name']} ({user['role']})")
    return jsonify({
        "access_token": token, "token_type": "bearer",
        "user": _user_out(user)
    })

@app.route("/api/auth/me")
@login_required
def auth_me():
    return jsonify(_user_out(g.user))

def _user_out(u):
    return {"id":u["id"],"username":u["username"],"full_name":u["full_name"],
            "email":u["email"],"role":u["role"],"is_active":bool(u["is_active"]),
            "client_id":u.get("client_id"), "created_at":u["created_at"]}


# ═══════════════════════════════════════════════════════════════
#  DASHBOARD
# ═══════════════════════════════════════════════════════════════
@app.route("/api/dashboard/")
@login_required
def dashboard():
    db = get_db()
    def cnt(sql, params=()): return db.execute(sql, params).fetchone()[0] or 0
    uid = g.user["id"]
    return jsonify({
        "total_clients": cnt("SELECT count(*) FROM clients"),
        "total_engagements": cnt("SELECT count(*) FROM engagements"),
        "total_tasks": cnt("SELECT count(*) FROM tasks"),
        "pending_tasks": cnt("SELECT count(*) FROM tasks WHERE status='Pending'"),
        "in_progress_tasks": cnt("SELECT count(*) FROM tasks WHERE status='In Progress'"),
        "completed_tasks": cnt("SELECT count(*) FROM tasks WHERE status='Completed'"),
        "under_review_tasks": cnt("SELECT count(*) FROM tasks WHERE status='Under Review'"),
        "approved_tasks": cnt("SELECT count(*) FROM tasks WHERE status='Approved'"),
        "rejected_tasks": cnt("SELECT count(*) FROM tasks WHERE status='Rejected'"),
        "open_queries": cnt("SELECT count(*) FROM queries WHERE status='Open'"),
        "total_queries": cnt("SELECT count(*) FROM queries"),
        "my_pending_tasks": cnt("SELECT count(*) FROM tasks WHERE (assignee_id=? OR id IN (SELECT task_id FROM task_assignees WHERE user_id=?)) AND status IN ('Pending','In Progress')", (uid, uid)),
        "my_review_tasks": cnt("SELECT count(*) FROM tasks WHERE status='Under Review'"),
    })


# ═══════════════════════════════════════════════════════════════
#  USERS
# ═══════════════════════════════════════════════════════════════
@app.route("/api/users/")
@login_required
def list_users():
    return jsonify([_user_out(r) for r in dict_rows(get_db().execute("SELECT * FROM users ORDER BY full_name").fetchall())])

@app.route("/api/users/", methods=["POST"])
@require_role("Admin")
def create_user():
    data = request.get_json(); db = get_db()
    if db.execute("SELECT id FROM users WHERE username=?", (data["username"],)).fetchone():
        return jsonify({"detail": "Username already exists"}), 400
    db.execute("INSERT INTO users (username, full_name, email, password_hash, role, client_id) VALUES (?,?,?,?,?,?)",
               (data["username"], data["full_name"], data.get("email"), hash_password(data["password"]), data.get("role","Member"), data.get("client_id")))
    db.commit()
    uid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    log_action(db, g.user["id"], "CREATE_USER", "User", uid, f"Created: {data['username']}", request.remote_addr)
    user = dict_row(db.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone())
    return jsonify(_user_out(user)), 201

@app.route("/api/users/<int:uid>", methods=["PUT"])
@require_role("Admin")
def update_user(uid):
    data = request.get_json(); db = get_db()
    fields, vals = [], []
    for f in ["full_name","email","role"]:
        if f in data: fields.append(f"{f}=?"); vals.append(data[f])
    if "is_active" in data: fields.append("is_active=?"); vals.append(1 if data["is_active"] else 0)
    if "client_id" in data: fields.append("client_id=?"); vals.append(data["client_id"])
    if not fields: return jsonify({"detail":"Nothing to update"}), 400
    vals.append(uid)
    db.execute(f"UPDATE users SET {','.join(fields)} WHERE id=?", vals); db.commit()
    log_action(db, g.user["id"], "UPDATE_USER", "User", uid, ip=request.remote_addr)
    user = dict_row(db.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone())
    return jsonify(_user_out(user))

@app.route("/api/users/<int:uid>/reset-password", methods=["POST"])
@require_role("Admin")
def reset_password(uid):
    db = get_db()
    db.execute("UPDATE users SET password_hash=? WHERE id=?", (hash_password("audit123"), uid)); db.commit()
    log_action(db, g.user["id"], "RESET_PASSWORD", "User", uid, ip=request.remote_addr)
    return jsonify({"message": "Password reset to 'audit123'"})

@app.route("/api/users/<int:uid>", methods=["DELETE"])
@require_role("Admin")
def delete_user(uid):
    db = get_db()
    if uid == g.user["id"]: return jsonify({"detail": "Cannot delete your own account"}), 400
    u = dict_row(db.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone())
    if not u: return jsonify({"detail": "Not found"}), 404
    log_action(db, g.user["id"], "DELETE_USER", "User", uid, f"Deleted: {u['username']}", request.remote_addr)
    db.execute("DELETE FROM users WHERE id=?", (uid,)); db.commit()
    return jsonify({"message": "User deleted"})

@app.route("/api/users/change-password", methods=["POST"])
@login_required
def change_password():
    data = request.get_json(); db = get_db()
    if not verify_password(data["old_password"], g.user["password_hash"]):
        return jsonify({"detail":"Current password incorrect"}), 400
    db.execute("UPDATE users SET password_hash=? WHERE id=?", (hash_password(data["new_password"]), g.user["id"])); db.commit()
    return jsonify({"message":"Password changed"})


# ═══════════════════════════════════════════════════════════════
#  CLIENTS
# ═══════════════════════════════════════════════════════════════
def _client_out(c):
    return {k: c[k] for k in ["id","name","pan","gstin","address","contact_person","contact_phone","contact_email","is_active","created_at"]}

@app.route("/api/clients/")
@login_required
def list_clients():
    return jsonify([_client_out(r) for r in dict_rows(get_db().execute("SELECT * FROM clients ORDER BY name").fetchall())])

@app.route("/api/clients/<int:cid>")
@login_required
def get_client(cid):
    c = dict_row(get_db().execute("SELECT * FROM clients WHERE id=?", (cid,)).fetchone())
    if not c: return jsonify({"detail":"Not found"}), 404
    return jsonify(_client_out(c))

@app.route("/api/clients/", methods=["POST"])
@login_required
def create_client():
    d = request.get_json(); db = get_db()
    db.execute("INSERT INTO clients (name,pan,gstin,address,contact_person,contact_phone,contact_email,created_by_id) VALUES (?,?,?,?,?,?,?,?)",
               (d["name"], d.get("pan"), d.get("gstin"), d.get("address"), d.get("contact_person"), d.get("contact_phone"), d.get("contact_email"), g.user["id"]))
    db.commit(); cid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    log_action(db, g.user["id"], "CREATE_CLIENT", "Client", cid, f"Created: {d['name']}", request.remote_addr)
    return jsonify(_client_out(dict_row(db.execute("SELECT * FROM clients WHERE id=?", (cid,)).fetchone()))), 201

@app.route("/api/clients/<int:cid>", methods=["PUT"])
@login_required
def update_client(cid):
    d = request.get_json(); db = get_db()
    fields, vals = [], []
    for f in ["name","pan","gstin","address","contact_person","contact_phone","contact_email"]:
        if f in d: fields.append(f"{f}=?"); vals.append(d[f])
    if "is_active" in d: fields.append("is_active=?"); vals.append(1 if d["is_active"] else 0)
    vals.append(cid)
    db.execute(f"UPDATE clients SET {','.join(fields)} WHERE id=?", vals); db.commit()
    log_action(db, g.user["id"], "UPDATE_CLIENT", "Client", cid, ip=request.remote_addr)
    return jsonify(_client_out(dict_row(db.execute("SELECT * FROM clients WHERE id=?", (cid,)).fetchone())))

@app.route("/api/clients/<int:cid>", methods=["DELETE"])
@require_role("Admin")
def delete_client(cid):
    db = get_db()
    c = dict_row(db.execute("SELECT * FROM clients WHERE id=?", (cid,)).fetchone())
    if not c: return jsonify({"detail": "Not found"}), 404
    log_action(db, g.user["id"], "DELETE_CLIENT", "Client", cid, f"Deleted: {c['name']}", request.remote_addr)
    db.execute("DELETE FROM clients WHERE id=?", (cid,)); db.commit()
    return jsonify({"message": "Client deleted"})


# ═══════════════════════════════════════════════════════════════
#  ENGAGEMENTS
# ═══════════════════════════════════════════════════════════════
def _eng_out(e, db):
    out = {k: e[k] for k in ["id","client_id","title","engagement_type","financial_year","period_from","period_to","team_leader_id","status","notes","created_at"]}
    cl = dict_row(db.execute("SELECT * FROM clients WHERE id=?", (e["client_id"],)).fetchone())
    out["client"] = _client_out(cl) if cl else None
    tl = dict_row(db.execute("SELECT * FROM users WHERE id=?", (e["team_leader_id"],)).fetchone()) if e["team_leader_id"] else None
    out["team_leader"] = _user_out(tl) if tl else None
    members = dict_rows(db.execute("""
        SELECT et.*, u.username, u.full_name, u.email, u.role, u.is_active, u.created_at as u_created
        FROM engagement_teams et JOIN users u ON et.user_id = u.id
        WHERE et.engagement_id=?""", (e["id"],)).fetchall())
    out["team_members"] = [{"id":m["id"],"user_id":m["user_id"],"role_in_engagement":m["role_in_engagement"],
        "user":{"id":m["user_id"],"username":m["username"],"full_name":m["full_name"],"email":m["email"],"role":m["role"],"is_active":bool(m["is_active"]),"created_at":m["u_created"]}} for m in members]
    return out

@app.route("/api/engagements/")
@login_required
def list_engagements():
    db = get_db()
    # Client role users can only see their own client's engagements
    if g.user["role"] == "Client":
        if not g.user.get("client_id"):
            return jsonify([])
        return jsonify([_eng_out(r, db) for r in dict_rows(db.execute(
            "SELECT * FROM engagements WHERE client_id=? ORDER BY created_at DESC",
            (g.user["client_id"],)).fetchall())])
    sql = "SELECT * FROM engagements"
    params = ()
    cid = request.args.get("client_id")
    if cid: sql += " WHERE client_id=?"; params = (cid,)
    sql += " ORDER BY created_at DESC"
    return jsonify([_eng_out(r, db) for r in dict_rows(db.execute(sql, params).fetchall())])

@app.route("/api/engagements/<int:eid>")
@login_required
def get_engagement(eid):
    db = get_db()
    e = dict_row(db.execute("SELECT * FROM engagements WHERE id=?", (eid,)).fetchone())
    if not e: return jsonify({"detail":"Not found"}), 404
    if g.user["role"] == "Client" and e["client_id"] != g.user.get("client_id"):
        return jsonify({"detail": "Access denied"}), 403
    return jsonify(_eng_out(e, db))

@app.route("/api/engagements/", methods=["POST"])
@login_required
def create_engagement():
    d = request.get_json(); db = get_db()
    db.execute("INSERT INTO engagements (client_id,title,engagement_type,financial_year,period_from,period_to,team_leader_id,notes,created_by_id) VALUES (?,?,?,?,?,?,?,?,?)",
               (d["client_id"], d["title"], d["engagement_type"], d["financial_year"], d.get("period_from"), d.get("period_to"), d.get("team_leader_id"), d.get("notes"), g.user["id"]))
    db.commit(); eid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    for uid in d.get("team_member_ids", []):
        db.execute("INSERT INTO engagement_teams (engagement_id, user_id) VALUES (?,?)", (eid, uid))
    db.commit()
    log_action(db, g.user["id"], "CREATE_ENGAGEMENT", "Engagement", eid, f"Created: {d['title']}", request.remote_addr)
    return jsonify(_eng_out(dict_row(db.execute("SELECT * FROM engagements WHERE id=?", (eid,)).fetchone()), db)), 201

@app.route("/api/engagements/<int:eid>", methods=["PUT"])
@login_required
def update_engagement(eid):
    d = request.get_json(); db = get_db()
    fields, vals = [], []
    for f in ["title","engagement_type","financial_year","period_from","period_to","team_leader_id","status","notes"]:
        if f in d: fields.append(f"{f}=?"); vals.append(d[f])
    vals.append(eid)
    db.execute(f"UPDATE engagements SET {','.join(fields)} WHERE id=?", vals); db.commit()
    log_action(db, g.user["id"], "UPDATE_ENGAGEMENT", "Engagement", eid, ip=request.remote_addr)
    return jsonify(_eng_out(dict_row(db.execute("SELECT * FROM engagements WHERE id=?", (eid,)).fetchone()), db))

@app.route("/api/engagements/<int:eid>", methods=["DELETE"])
@require_role("Admin", "Team Leader")
def delete_engagement(eid):
    db = get_db()
    e = dict_row(db.execute("SELECT * FROM engagements WHERE id=?", (eid,)).fetchone())
    if not e: return jsonify({"detail": "Not found"}), 404
    log_action(db, g.user["id"], "DELETE_ENGAGEMENT", "Engagement", eid, f"Deleted: {e['title']}", request.remote_addr)
    db.execute("DELETE FROM engagements WHERE id=?", (eid,)); db.commit()
    return jsonify({"message": "Engagement deleted"})

@app.route("/api/engagements/<int:eid>/add-member", methods=["POST"])
@login_required
def add_team_member(eid):
    uid = request.args.get("user_id"); db = get_db()
    if db.execute("SELECT id FROM engagement_teams WHERE engagement_id=? AND user_id=?", (eid, uid)).fetchone():
        return jsonify({"detail":"Already in team"}), 400
    db.execute("INSERT INTO engagement_teams (engagement_id, user_id, role_in_engagement) VALUES (?,?,?)",
               (eid, uid, request.args.get("role","Member"))); db.commit()
    return jsonify({"message":"Added"})

@app.route("/api/engagements/<int:eid>/apply-program/<int:pid>", methods=["POST"])
@login_required
def apply_program(eid, pid):
    db = get_db()
    prog = dict_row(db.execute("SELECT * FROM audit_programs WHERE id=?", (pid,)).fetchone())
    if not prog: return jsonify({"detail":"Program not found"}), 404
    items = dict_rows(db.execute("SELECT * FROM audit_checklist_items WHERE program_id=? ORDER BY sr_no", (pid,)).fetchall())
    for item in items:
        db.execute("INSERT INTO tasks (engagement_id,checklist_item_id,title,area,priority,status,working_paper_ref,created_by_id) VALUES (?,?,?,?,?,?,?,?)",
                   (eid, item["id"], item["description"], item["area"], item["priority"], "Pending", item["reference"], g.user["id"]))
    db.commit()
    log_action(db, g.user["id"], "APPLY_PROGRAM", "Engagement", eid, f"Applied '{prog['name']}', created {len(items)} tasks", request.remote_addr)
    return jsonify({"message": f"Created {len(items)} tasks from audit program '{prog['name']}'"})


# ═══════════════════════════════════════════════════════════════
#  AUDIT PROGRAMS
# ═══════════════════════════════════════════════════════════════
@app.route("/api/programs/")
@login_required
def list_programs():
    db = get_db()
    progs = dict_rows(db.execute("SELECT * FROM audit_programs WHERE is_active=1").fetchall())
    for p in progs:
        p["checklist_items"] = dict_rows(db.execute("SELECT * FROM audit_checklist_items WHERE program_id=? ORDER BY sr_no", (p["id"],)).fetchall())
    return jsonify(progs)

@app.route("/api/programs/<int:pid>")
@login_required
def get_program(pid):
    db = get_db()
    p = dict_row(db.execute("SELECT * FROM audit_programs WHERE id=?", (pid,)).fetchone())
    if not p: return jsonify({"detail":"Not found"}), 404
    p["checklist_items"] = dict_rows(db.execute("SELECT * FROM audit_checklist_items WHERE program_id=? ORDER BY sr_no", (pid,)).fetchall())
    return jsonify(p)

@app.route("/api/programs/", methods=["POST"])
@require_role("Admin","Team Leader")
def create_program():
    d = request.get_json(); db = get_db()
    db.execute("INSERT INTO audit_programs (name, engagement_type, description) VALUES (?,?,?)",
               (d["name"], d["engagement_type"], d.get("description")))
    db.commit(); pid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    for item in d.get("checklist_items", []):
        db.execute("INSERT INTO audit_checklist_items (program_id,sr_no,area,description,reference,priority) VALUES (?,?,?,?,?,?)",
                   (pid, item["sr_no"], item["area"], item["description"], item.get("reference"), item.get("priority","Medium")))
    db.commit()
    log_action(db, g.user["id"], "CREATE_PROGRAM", "AuditProgram", pid, f"Created: {d['name']}", request.remote_addr)
    return get_program(pid)

@app.route("/api/programs/<int:pid>", methods=["DELETE"])
@require_role("Admin", "Team Leader")
def delete_program(pid):
    db = get_db()
    p = dict_row(db.execute("SELECT * FROM audit_programs WHERE id=?", (pid,)).fetchone())
    if not p: return jsonify({"detail": "Not found"}), 404
    log_action(db, g.user["id"], "DELETE_PROGRAM", "AuditProgram", pid, f"Deleted: {p['name']}", request.remote_addr)
    db.execute("DELETE FROM audit_checklist_items WHERE program_id=?", (pid,))
    db.execute("DELETE FROM audit_programs WHERE id=?", (pid,)); db.commit()
    return jsonify({"message": "Program deleted"})


# ═══════════════════════════════════════════════════════════════
#  TASKS
# ═══════════════════════════════════════════════════════════════
def _task_out(t, db):
    out = {k: t[k] for k in ["id","engagement_id","checklist_item_id","title","description","area","assignee_id","status","priority","due_date","completed_at","working_paper_ref","created_at"]}
    # Primary assignee (legacy compat)
    if t["assignee_id"]:
        a = dict_row(db.execute("SELECT * FROM users WHERE id=?", (t["assignee_id"],)).fetchone())
        out["assignee"] = _user_out(a) if a else None
    else:
        out["assignee"] = None
    # All assignees from junction table
    rows = db.execute(
        "SELECT u.* FROM task_assignees ta JOIN users u ON u.id=ta.user_id WHERE ta.task_id=? ORDER BY u.full_name",
        (t["id"],)
    ).fetchall()
    out["assignees"] = [_user_out(dict_row(r)) for r in rows]
    out["assignee_ids"] = [r["id"] for r in [dict_row(x) for x in rows]]
    return out

def _sync_task_assignees(db, tid, user_ids):
    """Replace all assignees for a task; also keep legacy assignee_id as first assignee."""
    db.execute("DELETE FROM task_assignees WHERE task_id=?", (tid,))
    for uid in user_ids:
        try:
            db.execute("INSERT OR IGNORE INTO task_assignees (task_id, user_id) VALUES (?,?)", (tid, uid))
        except Exception:
            pass
    # Keep legacy assignee_id = first in list (or NULL)
    primary = user_ids[0] if user_ids else None
    db.execute("UPDATE tasks SET assignee_id=? WHERE id=?", (primary, tid))

@app.route("/api/tasks/")
@login_required
def list_tasks():
    db = get_db()
    sql = "SELECT * FROM tasks WHERE 1=1"
    params = []
    if request.args.get("engagement_id"):
        sql += " AND engagement_id=?"; params.append(request.args["engagement_id"])
    if request.args.get("assignee_id"):
        sql += " AND (assignee_id=? OR id IN (SELECT task_id FROM task_assignees WHERE user_id=?))"
        params += [request.args["assignee_id"], request.args["assignee_id"]]
    if request.args.get("status"):
        sql += " AND status=?"; params.append(request.args["status"])
    sql += " ORDER BY created_at DESC"
    return jsonify([_task_out(r, db) for r in dict_rows(db.execute(sql, params).fetchall())])

@app.route("/api/tasks/<int:tid>")
@login_required
def get_task(tid):
    db = get_db()
    t = dict_row(db.execute("SELECT * FROM tasks WHERE id=?", (tid,)).fetchone())
    if not t: return jsonify({"detail":"Not found"}), 404
    return jsonify(_task_out(t, db))

@app.route("/api/tasks/", methods=["POST"])
@login_required
def create_task():
    d = request.get_json(); db = get_db()
    # Support both single assignee_id and list of assignee_ids
    assignee_ids = d.get("assignee_ids") or ([d["assignee_id"]] if d.get("assignee_id") else [])
    primary_assignee = assignee_ids[0] if assignee_ids else None
    db.execute("INSERT INTO tasks (engagement_id,title,description,area,assignee_id,priority,due_date,working_paper_ref,created_by_id) VALUES (?,?,?,?,?,?,?,?,?)",
               (d["engagement_id"], d["title"], d.get("description"), d.get("area"), primary_assignee, d.get("priority","Medium"), d.get("due_date"), d.get("working_paper_ref"), g.user["id"]))
    db.commit(); tid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    _sync_task_assignees(db, tid, assignee_ids); db.commit()
    log_action(db, g.user["id"], "CREATE_TASK", "Task", tid, f"Created: {d['title']}", request.remote_addr)
    return jsonify(_task_out(dict_row(db.execute("SELECT * FROM tasks WHERE id=?", (tid,)).fetchone()), db)), 201

@app.route("/api/tasks/<int:tid>", methods=["PUT"])
@login_required
def update_task(tid):
    d = request.get_json(); db = get_db()
    fields, vals = [], []
    # Handle multi-assignee
    if "assignee_ids" in d:
        assignee_ids = d["assignee_ids"] or []
        primary = assignee_ids[0] if assignee_ids else None
        _sync_task_assignees(db, tid, assignee_ids)
        fields.append("assignee_id=?"); vals.append(primary)
    elif "assignee_id" in d:
        # Legacy single-assignee path
        aid = d["assignee_id"]
        _sync_task_assignees(db, tid, [aid] if aid else [])
        fields.append("assignee_id=?"); vals.append(aid)
    for f in ["title","description","area","status","priority","due_date","working_paper_ref"]:
        if f in d: fields.append(f"{f}=?"); vals.append(d[f])
    # Auto-set completed_at
    if d.get("status") == "Completed":
        cur = dict_row(db.execute("SELECT status FROM tasks WHERE id=?", (tid,)).fetchone())
        if cur and cur["status"] != "Completed":
            fields.append("completed_at=?"); vals.append(datetime.utcnow().isoformat())
    elif d.get("status") and d["status"] != "Completed":
        fields.append("completed_at=?"); vals.append(None)
    if fields:
        vals.append(tid)
        db.execute(f"UPDATE tasks SET {','.join(fields)} WHERE id=?", vals)
    db.commit()
    log_action(db, g.user["id"], "UPDATE_TASK", "Task", tid, f"Updated fields: {list(d.keys())}", request.remote_addr)
    return jsonify(_task_out(dict_row(db.execute("SELECT * FROM tasks WHERE id=?", (tid,)).fetchone()), db))

@app.route("/api/tasks/<int:tid>", methods=["DELETE"])
@login_required
def delete_task(tid):
    db = get_db()
    db.execute("DELETE FROM tasks WHERE id=?", (tid,)); db.commit()
    log_action(db, g.user["id"], "DELETE_TASK", "Task", tid, ip=request.remote_addr)
    return jsonify({"message":"Deleted"})


# ═══════════════════════════════════════════════════════════════
#  COMMENTS
# ═══════════════════════════════════════════════════════════════
def _comment_out(c, db):
    out = {k: c[k] for k in ["id","task_id","author_id","content","is_query","created_at","updated_at"]}
    out["is_query"] = bool(c["is_query"])
    a = dict_row(db.execute("SELECT * FROM users WHERE id=?", (c["author_id"],)).fetchone())
    out["author"] = _user_out(a) if a else None
    return out

@app.route("/api/comments/")
@login_required
def list_comments():
    db = get_db()
    sql = "SELECT * FROM comments"
    params = ()
    if request.args.get("task_id"):
        sql += " WHERE task_id=?"; params = (request.args["task_id"],)
    sql += " ORDER BY created_at DESC"
    return jsonify([_comment_out(r, db) for r in dict_rows(db.execute(sql, params).fetchall())])

@app.route("/api/comments/", methods=["POST"])
@login_required
def create_comment():
    d = request.get_json(); db = get_db()
    is_q = 1 if d.get("is_query") else 0
    db.execute("INSERT INTO comments (task_id, author_id, content, is_query) VALUES (?,?,?,?)",
               (d["task_id"], g.user["id"], d["content"], is_q))
    db.commit(); cid = db.execute("SELECT last_insert_rowid()").fetchone()[0]

    if is_q:
        task = dict_row(db.execute("SELECT * FROM tasks WHERE id=?", (d["task_id"],)).fetchone())
        if task:
            max_sr = db.execute("SELECT COALESCE(MAX(sr_no),0) FROM queries WHERE engagement_id=?", (task["engagement_id"],)).fetchone()[0]
            db.execute("INSERT INTO queries (engagement_id, comment_id, sr_no, query_text, raised_by_id, task_reference) VALUES (?,?,?,?,?,?)",
                       (task["engagement_id"], cid, max_sr+1, d["content"], g.user["id"], f"Task #{task['id']}: {task['title']}"))
            db.commit()

    log_action(db, g.user["id"], "CREATE_COMMENT", "Comment", cid, f"Task #{d['task_id']}" + (" [QUERY]" if is_q else ""), request.remote_addr)
    return jsonify(_comment_out(dict_row(db.execute("SELECT * FROM comments WHERE id=?", (cid,)).fetchone()), db)), 201

@app.route("/api/comments/<int:cid>", methods=["PUT"])
@login_required
def update_comment(cid):
    d = request.get_json(); db = get_db()
    comment = dict_row(db.execute("SELECT * FROM comments WHERE id=?", (cid,)).fetchone())
    if not comment: return jsonify({"detail":"Not found"}), 404
    if comment["author_id"] != g.user["id"]:
        return jsonify({"detail":"Can only edit own comments"}), 403

    was_query = comment["is_query"]
    fields, vals = ["updated_at=?"], [datetime.utcnow().isoformat()]
    if "content" in d: fields.append("content=?"); vals.append(d["content"])
    if "is_query" in d:
        new_q = 1 if d["is_query"] else 0
        fields.append("is_query=?"); vals.append(new_q)
        # Handle query flag changes
        if new_q and not was_query:
            task = dict_row(db.execute("SELECT * FROM tasks WHERE id=?", (comment["task_id"],)).fetchone())
            if task:
                max_sr = db.execute("SELECT COALESCE(MAX(sr_no),0) FROM queries WHERE engagement_id=?", (task["engagement_id"],)).fetchone()[0]
                db.execute("INSERT INTO queries (engagement_id, comment_id, sr_no, query_text, raised_by_id, task_reference) VALUES (?,?,?,?,?,?)",
                           (task["engagement_id"], cid, max_sr+1, d.get("content", comment["content"]), g.user["id"], f"Task #{task['id']}: {task['title']}"))
        elif not new_q and was_query:
            db.execute("DELETE FROM queries WHERE comment_id=?", (cid,))

    vals.append(cid)
    db.execute(f"UPDATE comments SET {','.join(fields)} WHERE id=?", vals); db.commit()
    log_action(db, g.user["id"], "UPDATE_COMMENT", "Comment", cid, ip=request.remote_addr)
    return jsonify(_comment_out(dict_row(db.execute("SELECT * FROM comments WHERE id=?", (cid,)).fetchone()), db))

@app.route("/api/comments/<int:cid>", methods=["DELETE"])
@login_required
def delete_comment(cid):
    db = get_db()
    comment = dict_row(db.execute("SELECT * FROM comments WHERE id=?", (cid,)).fetchone())
    if not comment: return jsonify({"detail":"Not found"}), 404
    if comment["author_id"] != g.user["id"]:
        return jsonify({"detail":"Can only delete own comments"}), 403
    db.execute("DELETE FROM queries WHERE comment_id=?", (cid,))
    db.execute("DELETE FROM comments WHERE id=?", (cid,)); db.commit()
    log_action(db, g.user["id"], "DELETE_COMMENT", "Comment", cid, ip=request.remote_addr)
    return jsonify({"message":"Deleted"})


# ═══════════════════════════════════════════════════════════════
#  QUERIES
# ═══════════════════════════════════════════════════════════════
def _query_out(q, db):
    out = {k: q[k] for k in ["id","engagement_id","comment_id","sr_no","query_text","response","status","raised_by_id","raised_date","responded_by_id","responded_date","task_reference"]}
    rb = dict_row(db.execute("SELECT * FROM users WHERE id=?", (q["raised_by_id"],)).fetchone())
    out["raised_by"] = _user_out(rb) if rb else None
    if q["responded_by_id"]:
        rsp = dict_row(db.execute("SELECT * FROM users WHERE id=?", (q["responded_by_id"],)).fetchone())
        out["responded_by"] = _user_out(rsp) if rsp else None
    else:
        out["responded_by"] = None
    return out

@app.route("/api/queries/")
@login_required
def list_queries():
    db = get_db()
    sql = "SELECT * FROM queries WHERE 1=1"
    params = []
    if request.args.get("engagement_id"):
        sql += " AND engagement_id=?"; params.append(request.args["engagement_id"])
    if request.args.get("status"):
        sql += " AND status=?"; params.append(request.args["status"])
    sql += " ORDER BY sr_no"
    return jsonify([_query_out(r, db) for r in dict_rows(db.execute(sql, params).fetchall())])

@app.route("/api/queries/", methods=["POST"])
@login_required
def create_manual_query():
    d = request.get_json(); db = get_db()
    max_sr = db.execute("SELECT COALESCE(MAX(sr_no),0) FROM queries WHERE engagement_id=?", (d["engagement_id"],)).fetchone()[0]
    db.execute("INSERT INTO queries (engagement_id, sr_no, query_text, raised_by_id, task_reference) VALUES (?,?,?,?,?)",
               (d["engagement_id"], max_sr+1, d["query_text"], g.user["id"], d.get("task_reference")))
    db.commit(); qid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    log_action(db, g.user["id"], "CREATE_QUERY", "Query", qid, f"Manual query", request.remote_addr)
    return jsonify(_query_out(dict_row(db.execute("SELECT * FROM queries WHERE id=?", (qid,)).fetchone()), db)), 201

@app.route("/api/queries/<int:qid>", methods=["PUT"])
@login_required
def update_query(qid):
    d = request.get_json(); db = get_db()
    fields, vals = [], []
    if "response" in d:
        fields += ["response=?", "responded_by_id=?", "responded_date=?"]
        vals += [d["response"], g.user["id"], datetime.utcnow().isoformat()]
        if "status" not in d: fields.append("status=?"); vals.append("Responded")
    if "status" in d: fields.append("status=?"); vals.append(d["status"])
    vals.append(qid)
    db.execute(f"UPDATE queries SET {','.join(fields)} WHERE id=?", vals); db.commit()
    log_action(db, g.user["id"], "UPDATE_QUERY", "Query", qid, ip=request.remote_addr)
    return jsonify(_query_out(dict_row(db.execute("SELECT * FROM queries WHERE id=?", (qid,)).fetchone()), db))

@app.route("/api/queries/<int:qid>", methods=["DELETE"])
@require_role("Admin", "Team Leader")
def delete_query(qid):
    db = get_db()
    q = dict_row(db.execute("SELECT * FROM queries WHERE id=?", (qid,)).fetchone())
    if not q: return jsonify({"detail": "Not found"}), 404
    log_action(db, g.user["id"], "DELETE_QUERY", "Query", qid, ip=request.remote_addr)
    db.execute("DELETE FROM queries WHERE id=?", (qid,)); db.commit()
    return jsonify({"message": "Query deleted"})

@app.route("/api/queries/export/<int:eid>")
@login_required
def export_queries(eid):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    db = get_db()
    queries = dict_rows(db.execute("SELECT q.*, u1.full_name as raised_name, u2.full_name as resp_name FROM queries q LEFT JOIN users u1 ON q.raised_by_id=u1.id LEFT JOIN users u2 ON q.responded_by_id=u2.id WHERE q.engagement_id=? ORDER BY q.sr_no", (eid,)).fetchall())

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Query Sheet"
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill("solid", fgColor="1B3A5C")
    data_font = Font(name="Arial", size=10)
    border = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    wrap = Alignment(wrap_text=True, vertical="top")
    alt = PatternFill("solid", fgColor="F0F4F8")

    # Firm stamp at top
    start_row = 1
    if FIRM_NAME:
        firm_cell = ws.cell(row=1, column=1, value=f"{FIRM_NAME}  |  Reg. No.: {FIRM_REG_NO}  |  Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}")
        firm_cell.font = Font(name="Arial", size=9, italic=True, color="5A6472")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
        start_row = 2

    headers = ["Sr. No.","Query","Response","Status","Raised By","Raised Date","Responded By","Response Date","Task Reference"]
    widths = [8,40,40,12,15,14,15,14,25]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=start_row, column=col, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    for ri, q in enumerate(queries, start_row + 1):
        vals = [q["sr_no"], q["query_text"], q.get("response",""), q["status"], q.get("raised_name",""),
                q["raised_date"][:10] if q.get("raised_date") else "", q.get("resp_name",""),
                q["responded_date"][:10] if q.get("responded_date") else "", q.get("task_reference","")]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = data_font; c.alignment = wrap; c.border = border
            if (ri - start_row) % 2 == 0: c.fill = alt

    ws.auto_filter.ref = f"A{start_row}:I{start_row}"; ws.freeze_panes = f"A{start_row + 1}"
    fn = f"Query_Sheet_Engagement_{eid}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    fp = str(EXPORT_DIR / fn); wb.save(fp)
    return send_file(fp, as_attachment=True, download_name=fn, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ═══════════════════════════════════════════════════════════════
#  FILE UPLOADS
# ═══════════════════════════════════════════════════════════════
def _file_out(f, db):
    out = {k: f[k] for k in ["id","task_id","filename","original_filename","file_size","mime_type","uploaded_by_id","uploaded_at"]}
    u = dict_row(db.execute("SELECT * FROM users WHERE id=?", (f["uploaded_by_id"],)).fetchone())
    out["uploaded_by"] = _user_out(u) if u else None
    return out

@app.route("/api/files/")
@login_required
def list_files():
    db = get_db()
    tid = request.args.get("task_id")
    return jsonify([_file_out(r, db) for r in dict_rows(db.execute("SELECT * FROM file_uploads WHERE task_id=? ORDER BY uploaded_at DESC", (tid,)).fetchall())])

@app.route("/api/files/upload", methods=["POST"])
@login_required
def upload_file():
    db = get_db()
    tid = request.form.get("task_id")
    f = request.files.get("file")
    if not f: return jsonify({"detail":"No file"}), 400

    task = dict_row(db.execute("SELECT * FROM tasks WHERE id=?", (tid,)).fetchone())
    if not task: return jsonify({"detail":"Task not found"}), 404

    folder = UPLOAD_DIR / f"engagement_{task['engagement_id']}" / f"task_{tid}"
    folder.mkdir(parents=True, exist_ok=True)
    safe = f"{uuid.uuid4().hex[:12]}_{f.filename}"
    fp = folder / safe
    content = f.read(); fp.write_bytes(content)

    db.execute("INSERT INTO file_uploads (task_id,filename,original_filename,file_path,file_size,mime_type,uploaded_by_id) VALUES (?,?,?,?,?,?,?)",
               (tid, safe, f.filename, str(fp), len(content), f.content_type, g.user["id"]))
    db.commit(); fid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    log_action(db, g.user["id"], "UPLOAD_FILE", "FileUpload", fid, f"Uploaded: {f.filename}", request.remote_addr)
    return jsonify(_file_out(dict_row(db.execute("SELECT * FROM file_uploads WHERE id=?", (fid,)).fetchone()), db)), 201

@app.route("/api/files/download/<int:fid>")
@login_required
def download_file(fid):
    db = get_db()
    f = dict_row(db.execute("SELECT * FROM file_uploads WHERE id=?", (fid,)).fetchone())
    if not f: return jsonify({"detail":"Not found"}), 404
    return send_file(f["file_path"], as_attachment=True, download_name=f["original_filename"])

@app.route("/api/files/<int:fid>", methods=["DELETE"])
@login_required
def delete_file(fid):
    db = get_db()
    f = dict_row(db.execute("SELECT * FROM file_uploads WHERE id=?", (fid,)).fetchone())
    if not f: return jsonify({"detail":"Not found"}), 404
    fp = Path(f["file_path"])
    if fp.exists(): fp.unlink()
    db.execute("DELETE FROM file_uploads WHERE id=?", (fid,)); db.commit()
    log_action(db, g.user["id"], "DELETE_FILE", "FileUpload", fid, f"Deleted: {f['original_filename']}", request.remote_addr)
    return jsonify({"message":"Deleted"})


# ═══════════════════════════════════════════════════════════════
#  REVIEWS
# ═══════════════════════════════════════════════════════════════
@app.route("/api/reviews/")
@login_required
def list_reviews():
    db = get_db()
    tid = request.args.get("task_id")
    rows = dict_rows(db.execute("SELECT r.*, u.full_name, u.username, u.email, u.role, u.is_active, u.created_at as u_created FROM reviews r JOIN users u ON r.reviewer_id=u.id WHERE r.task_id=? ORDER BY r.reviewed_at DESC", (tid,)).fetchall())
    return jsonify([{"id":r["id"],"task_id":r["task_id"],"reviewer_id":r["reviewer_id"],"action":r["action"],"remarks":r["remarks"],"reviewed_at":r["reviewed_at"],
        "reviewer":{"id":r["reviewer_id"],"username":r["username"],"full_name":r["full_name"],"email":r["email"],"role":r["role"],"is_active":bool(r["is_active"]),"created_at":r["u_created"]}} for r in rows])

@app.route("/api/reviews/", methods=["POST"])
@require_role("Admin","Team Leader")
def submit_review():
    d = request.get_json(); db = get_db()
    task = dict_row(db.execute("SELECT * FROM tasks WHERE id=?", (d["task_id"],)).fetchone())
    if not task: return jsonify({"detail":"Task not found"}), 404
    if task["status"] not in ("Under Review","Completed"):
        return jsonify({"detail":"Task must be Under Review or Completed"}), 400

    db.execute("INSERT INTO reviews (task_id, reviewer_id, action, remarks) VALUES (?,?,?,?)",
               (d["task_id"], g.user["id"], d["action"], d.get("remarks")))
    new_status = "Approved" if d["action"] == "Approved" else "Rejected"
    db.execute("UPDATE tasks SET status=? WHERE id=?", (new_status, d["task_id"]))
    db.commit()
    rid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    log_action(db, g.user["id"], "REVIEW_TASK", "Review", rid, f"Task #{d['task_id']}: {d['action']}", request.remote_addr)

    row = dict_row(db.execute("SELECT r.*, u.full_name, u.username, u.email, u.role, u.is_active, u.created_at as u_created FROM reviews r JOIN users u ON r.reviewer_id=u.id WHERE r.id=?", (rid,)).fetchone())
    return jsonify({"id":row["id"],"task_id":row["task_id"],"reviewer_id":row["reviewer_id"],"action":row["action"],"remarks":row["remarks"],"reviewed_at":row["reviewed_at"],
        "reviewer":{"id":row["reviewer_id"],"username":row["username"],"full_name":row["full_name"],"email":row["email"],"role":row["role"],"is_active":bool(row["is_active"]),"created_at":row["u_created"]}})


# ═══════════════════════════════════════════════════════════════
#  AUDIT LOGS
# ═══════════════════════════════════════════════════════════════
@app.route("/api/audit-logs/")
@require_role("Admin","Team Leader")
def list_audit_logs():
    db = get_db()
    limit = request.args.get("limit", 200, type=int)
    rows = dict_rows(db.execute("""
        SELECT al.*, u.full_name, u.username FROM audit_logs al
        LEFT JOIN users u ON al.user_id = u.id
        ORDER BY al.timestamp DESC LIMIT ?""", (limit,)).fetchall())
    return jsonify([{
        "id":r["id"],"user_id":r["user_id"],"action":r["action"],"entity_type":r["entity_type"],
        "entity_id":r["entity_id"],"details":r["details"],"ip_address":r["ip_address"],"timestamp":r["timestamp"],
        "user":{"id":r["user_id"],"username":r.get("username",""),"full_name":r.get("full_name","System"),"email":None,"role":"","is_active":True,"created_at":""} if r["user_id"] else None
    } for r in rows])

@app.route("/api/audit-logs/<int:lid>", methods=["DELETE"])
@require_role("Admin")
def delete_audit_log(lid):
    db = get_db()
    if not db.execute("SELECT id FROM audit_logs WHERE id=?", (lid,)).fetchone():
        return jsonify({"detail": "Not found"}), 404
    db.execute("DELETE FROM audit_logs WHERE id=?", (lid,)); db.commit()
    return jsonify({"message": "Log deleted"})

@app.route("/api/audit-logs/clear", methods=["DELETE"])
@require_role("Admin")
def clear_audit_logs():
    db = get_db()
    db.execute("DELETE FROM audit_logs"); db.commit()
    return jsonify({"message": "All audit logs cleared"})


# ═══════════════════════════════════════════════════════════════
#  AUDIT BOOKLET
# ═══════════════════════════════════════════════════════════════
@app.route("/api/booklet/generate/<int:eid>")
@login_required
def gen_booklet(eid):
    db = get_db()
    eng = dict_row(db.execute("SELECT e.*, c.name as client_name, c.pan, c.gstin FROM engagements e JOIN clients c ON e.client_id=c.id WHERE e.id=?", (eid,)).fetchone())
    if not eng: return jsonify({"detail":"Not found"}), 404

    # Team — combine: team leader + engagement_teams + anyone assigned to any task in this engagement
    seen_ids = set()
    team = []
    if eng["team_leader_id"]:
        tl = dict_row(db.execute("SELECT * FROM users WHERE id=?", (eng["team_leader_id"],)).fetchone())
        if tl:
            team.append({"full_name": tl["full_name"], "role": "Team Leader", "email": tl["email"] or ""})
            seen_ids.add(tl["id"])
    for m in dict_rows(db.execute("SELECT u.* FROM engagement_teams et JOIN users u ON et.user_id=u.id WHERE et.engagement_id=?", (eid,)).fetchall()):
        if m["id"] not in seen_ids:
            team.append({"full_name": m["full_name"], "role": m.get("role","Member"), "email": m["email"] or ""})
            seen_ids.add(m["id"])
    # Also pull in every user assigned to any task in this engagement via task_assignees
    for m in dict_rows(db.execute(
        """SELECT DISTINCT u.* FROM task_assignees ta
           JOIN tasks t ON ta.task_id=t.id
           JOIN users u ON ta.user_id=u.id
           WHERE t.engagement_id=? ORDER BY u.full_name""", (eid,)).fetchall()):
        if m["id"] not in seen_ids:
            team.append({"full_name": m["full_name"], "role": m.get("role","Audit Member"), "email": m["email"] or ""})
            seen_ids.add(m["id"])

    # Tasks — collect all assignee names from task_assignees junction table
    tasks_raw = dict_rows(db.execute(
        "SELECT * FROM tasks WHERE engagement_id=? ORDER BY id", (eid,)).fetchall())
    # Build map: task_id -> comma-separated assignee names
    task_assignee_names = {}
    if tasks_raw:
        tids_all = [t["id"] for t in tasks_raw]
        ph_all = ",".join("?" * len(tids_all))
        for row in dict_rows(db.execute(
            f"""SELECT ta.task_id, u.full_name FROM task_assignees ta
                JOIN users u ON ta.user_id=u.id
                WHERE ta.task_id IN ({ph_all}) ORDER BY u.full_name""", tids_all).fetchall()):
            task_assignee_names.setdefault(row["task_id"], []).append(row["full_name"])
        # Fallback: also check legacy assignee_id for tasks not yet in junction table
        for t in tasks_raw:
            if t["id"] not in task_assignee_names and t.get("assignee_id"):
                u = dict_row(db.execute("SELECT full_name FROM users WHERE id=?", (t["assignee_id"],)).fetchone())
                if u: task_assignee_names[t["id"]] = [u["full_name"]]
    tasks = [{"id": t["id"], "title": t["title"], "area": t.get("area",""), "status": t["status"],
              "priority": t.get("priority",""),
              "assignee_name": ", ".join(task_assignee_names.get(t["id"], [])) or "Unassigned",
              "working_paper_ref": t.get("working_paper_ref","")} for t in tasks_raw]
    tids = [t["id"] for t in tasks]

    # Comments
    cbt = {}
    if tids:
        ph = ",".join("?" * len(tids))
        for c in dict_rows(db.execute(f"SELECT c.*, u.full_name as author_name FROM comments c JOIN users u ON c.author_id=u.id WHERE c.task_id IN ({ph}) ORDER BY c.created_at", tids).fetchall()):
            cbt.setdefault(c["task_id"],[]).append({"author_name":c["author_name"],"content":c["content"],"is_query":bool(c["is_query"]),"created_at":c["created_at"][:16] if c["created_at"] else ""})

    # Reviews
    rbt = {}
    if tids:
        for r in dict_rows(db.execute(f"SELECT r.*, u.full_name as reviewer_name FROM reviews r JOIN users u ON r.reviewer_id=u.id WHERE r.task_id IN ({ph}) ORDER BY r.reviewed_at", tids).fetchall()):
            rbt.setdefault(r["task_id"],[]).append({"reviewer_name":r["reviewer_name"],"action":r["action"],"remarks":r.get("remarks",""),"reviewed_at":r["reviewed_at"][:16] if r["reviewed_at"] else ""})

    # Queries
    qs = [{"sr_no":q["sr_no"],"query_text":q["query_text"],"response":q.get("response",""),"status":q["status"],
           "raised_by_name":q.get("raised_name",""),"raised_date":q["raised_date"][:10] if q.get("raised_date") else "",
           "task_reference":q.get("task_reference","")}
        for q in dict_rows(db.execute("SELECT q.*, u.full_name as raised_name FROM queries q LEFT JOIN users u ON q.raised_by_id=u.id WHERE q.engagement_id=? ORDER BY q.sr_no", (eid,)).fetchall())]

    fn = f"Audit_Booklet_{eng['client_name']}_{eng['financial_year']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.docx".replace(" ","_")
    fp = str(BOOKLET_DIR / fn)
    generate_booklet(eng, tasks, cbt, rbt, qs, team, fp,
                     firm_name=FIRM_NAME, firm_reg_no=FIRM_REG_NO)
    log_action(db, g.user["id"], "GENERATE_BOOKLET", "Engagement", eid, f"Generated: {fn}", request.remote_addr)
    return send_file(fp, as_attachment=True, download_name=fn, mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document")


# ═══════════════════════════════════════════════════════════════
#  TIME TRACKING
# ═══════════════════════════════════════════════════════════════
def _tlog_out(t, db):
    out = {k: t[k] for k in ["id","task_id","user_id","date","hours","note","created_at"]}
    u = dict_row(db.execute("SELECT * FROM users WHERE id=?", (t["user_id"],)).fetchone())
    out["user"] = _user_out(u) if u else None
    return out

@app.route("/api/time-logs/")
@login_required
def list_time_logs():
    db = get_db()
    sql = "SELECT * FROM time_logs WHERE 1=1"
    params = []
    if request.args.get("task_id"):
        sql += " AND task_id=?"; params.append(request.args["task_id"])
    if request.args.get("user_id"):
        sql += " AND user_id=?"; params.append(request.args["user_id"])
    if request.args.get("engagement_id"):
        sql += """ AND task_id IN (SELECT id FROM tasks WHERE engagement_id=?)"""
        params.append(request.args["engagement_id"])
    sql += " ORDER BY date DESC, created_at DESC"
    return jsonify([_tlog_out(r, db) for r in dict_rows(db.execute(sql, params).fetchall())])

@app.route("/api/time-logs/", methods=["POST"])
@login_required
def create_time_log():
    d = request.get_json(); db = get_db()
    hours = float(d.get("hours", 0))
    if hours <= 0:
        return jsonify({"detail": "Hours must be > 0"}), 400
    db.execute("INSERT INTO time_logs (task_id, user_id, date, hours, note) VALUES (?,?,?,?,?)",
               (d["task_id"], g.user["id"], d.get("date", datetime.utcnow().strftime("%Y-%m-%d")), hours, d.get("note")))
    db.commit(); lid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    log_action(db, g.user["id"], "LOG_TIME", "Task", d["task_id"], f"{hours}h on task #{d['task_id']}", request.remote_addr)
    return jsonify(_tlog_out(dict_row(db.execute("SELECT * FROM time_logs WHERE id=?", (lid,)).fetchone()), db)), 201

@app.route("/api/time-logs/<int:lid>", methods=["DELETE"])
@login_required
def delete_time_log(lid):
    db = get_db()
    tl = dict_row(db.execute("SELECT * FROM time_logs WHERE id=?", (lid,)).fetchone())
    if not tl: return jsonify({"detail":"Not found"}), 404
    if tl["user_id"] != g.user["id"] and g.user["role"] not in ("Admin","Team Leader"):
        return jsonify({"detail":"Access denied"}), 403
    db.execute("DELETE FROM time_logs WHERE id=?", (lid,)); db.commit()
    return jsonify({"message":"Deleted"})

@app.route("/api/time-logs/summary/")
@login_required
def time_log_summary():
    db = get_db()
    eid = request.args.get("engagement_id")
    if eid:
        rows = dict_rows(db.execute("""
            SELECT tl.user_id, u.full_name, SUM(tl.hours) as total_hours, COUNT(tl.id) as entries,
                   t.engagement_id
            FROM time_logs tl
            JOIN tasks t ON tl.task_id = t.id
            JOIN users u ON tl.user_id = u.id
            WHERE t.engagement_id=?
            GROUP BY tl.user_id, u.full_name""", (eid,)).fetchall())
        return jsonify([dict(r) for r in rows])
    rows = dict_rows(db.execute("""
        SELECT tl.user_id, u.full_name, SUM(tl.hours) as total_hours, COUNT(tl.id) as entries
        FROM time_logs tl JOIN users u ON tl.user_id=u.id
        GROUP BY tl.user_id, u.full_name ORDER BY total_hours DESC""").fetchall())
    return jsonify([dict(r) for r in rows])


# ═══════════════════════════════════════════════════════════════
#  INVOICES / FEE MANAGEMENT
# ═══════════════════════════════════════════════════════════════
def _inv_out(inv, db):
    out = {k: inv[k] for k in ["id","engagement_id","client_id","invoice_no","invoice_date","description","amount","gst_percent","gst_amount","total_amount","payment_status","payment_date","payment_note","created_at"]}
    cl = dict_row(db.execute("SELECT * FROM clients WHERE id=?", (inv["client_id"],)).fetchone())
    out["client"] = _client_out(cl) if cl else None
    eng = dict_row(db.execute("SELECT * FROM engagements WHERE id=?", (inv["engagement_id"],)).fetchone())
    out["engagement_title"] = eng["title"] if eng else ""
    out["engagement_fy"] = eng["financial_year"] if eng else ""
    return out

@app.route("/api/invoices/")
@login_required
def list_invoices():
    db = get_db()
    sql = "SELECT * FROM invoices WHERE 1=1"
    params = []
    if request.args.get("client_id"):
        sql += " AND client_id=?"; params.append(request.args["client_id"])
    if request.args.get("engagement_id"):
        sql += " AND engagement_id=?"; params.append(request.args["engagement_id"])
    if request.args.get("payment_status"):
        sql += " AND payment_status=?"; params.append(request.args["payment_status"])
    sql += " ORDER BY invoice_date DESC"
    return jsonify([_inv_out(r, db) for r in dict_rows(db.execute(sql, params).fetchall())])

@app.route("/api/invoices/", methods=["POST"])
@require_role("Admin", "Team Leader")
def create_invoice():
    d = request.get_json(); db = get_db()
    amount = float(d.get("amount", 0))
    gst_pct = float(d.get("gst_percent", 18.0))
    gst_amount = round(amount * gst_pct / 100, 2)
    total = round(amount + gst_amount, 2)
    # Auto-generate invoice no if not provided
    inv_no = d.get("invoice_no") or f"INV/{datetime.now().strftime('%y%m')}/{db.execute('SELECT COALESCE(MAX(id),0)+1 FROM invoices').fetchone()[0]:04d}"
    db.execute("INSERT INTO invoices (engagement_id,client_id,invoice_no,invoice_date,description,amount,gst_percent,gst_amount,total_amount,payment_status,created_by_id) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
               (d["engagement_id"], d["client_id"], inv_no, d["invoice_date"], d.get("description"), amount, gst_pct, gst_amount, total, d.get("payment_status","Unpaid"), g.user["id"]))
    db.commit(); iid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    log_action(db, g.user["id"], "CREATE_INVOICE", "Invoice", iid, f"Invoice {inv_no} for ₹{total}", request.remote_addr)
    return jsonify(_inv_out(dict_row(db.execute("SELECT * FROM invoices WHERE id=?", (iid,)).fetchone()), db)), 201

@app.route("/api/invoices/<int:iid>", methods=["PUT"])
@require_role("Admin", "Team Leader")
def update_invoice(iid):
    d = request.get_json(); db = get_db()
    fields, vals = [], []
    if "amount" in d or "gst_percent" in d:
        inv = dict_row(db.execute("SELECT * FROM invoices WHERE id=?", (iid,)).fetchone())
        amount = float(d.get("amount", inv["amount"]))
        gst_pct = float(d.get("gst_percent", inv["gst_percent"]))
        gst_amount = round(amount * gst_pct / 100, 2)
        total = round(amount + gst_amount, 2)
        fields += ["amount=?","gst_percent=?","gst_amount=?","total_amount=?"]; vals += [amount, gst_pct, gst_amount, total]
    for f in ["invoice_date","description","payment_status","payment_date","payment_note"]:
        if f in d: fields.append(f"{f}=?"); vals.append(d[f])
    if not fields: return jsonify({"detail":"Nothing to update"}), 400
    vals.append(iid)
    db.execute(f"UPDATE invoices SET {','.join(fields)} WHERE id=?", vals); db.commit()
    log_action(db, g.user["id"], "UPDATE_INVOICE", "Invoice", iid, ip=request.remote_addr)
    return jsonify(_inv_out(dict_row(db.execute("SELECT * FROM invoices WHERE id=?", (iid,)).fetchone()), db))

@app.route("/api/invoices/<int:iid>", methods=["DELETE"])
@require_role("Admin")
def delete_invoice(iid):
    db = get_db()
    db.execute("DELETE FROM invoices WHERE id=?", (iid,)); db.commit()
    log_action(db, g.user["id"], "DELETE_INVOICE", "Invoice", iid, ip=request.remote_addr)
    return jsonify({"message":"Deleted"})

@app.route("/api/invoices/summary/")
@login_required
def invoice_summary():
    db = get_db()
    row = dict_row(db.execute("""SELECT
        COALESCE(SUM(total_amount),0) as total_billed,
        COALESCE(SUM(CASE WHEN payment_status='Paid' THEN total_amount ELSE 0 END),0) as total_received,
        COALESCE(SUM(CASE WHEN payment_status='Unpaid' THEN total_amount ELSE 0 END),0) as total_outstanding,
        COUNT(*) as total_invoices,
        SUM(CASE WHEN payment_status='Unpaid' THEN 1 ELSE 0 END) as unpaid_count
        FROM invoices""").fetchone())
    return jsonify(dict(row))


# ═══════════════════════════════════════════════════════════════
#  DOCUMENT INWARD / OUTWARD REGISTER
# ═══════════════════════════════════════════════════════════════
def _doc_out(d, db):
    out = {k: d[k] for k in ["id","engagement_id","client_id","doc_type","doc_name","doc_category","doc_date","received_from","sent_to","reference_no","status","remarks","created_at"]}
    if d.get("client_id"):
        cl = dict_row(db.execute("SELECT * FROM clients WHERE id=?", (d["client_id"],)).fetchone())
        out["client_name"] = cl["name"] if cl else ""
    else:
        out["client_name"] = ""
    if d.get("engagement_id"):
        eng = dict_row(db.execute("SELECT * FROM engagements WHERE id=?", (d["engagement_id"],)).fetchone())
        out["engagement_title"] = eng["title"] if eng else ""
    else:
        out["engagement_title"] = ""
    return out

@app.route("/api/doc-register/")
@login_required
def list_doc_register():
    db = get_db()
    sql = "SELECT * FROM doc_register WHERE 1=1"
    params = []
    if request.args.get("client_id"):
        sql += " AND client_id=?"; params.append(request.args["client_id"])
    if request.args.get("engagement_id"):
        sql += " AND engagement_id=?"; params.append(request.args["engagement_id"])
    if request.args.get("doc_type"):
        sql += " AND doc_type=?"; params.append(request.args["doc_type"])
    sql += " ORDER BY doc_date DESC, created_at DESC"
    return jsonify([_doc_out(r, db) for r in dict_rows(db.execute(sql, params).fetchall())])

@app.route("/api/doc-register/", methods=["POST"])
@login_required
def create_doc_entry():
    d = request.get_json(); db = get_db()
    db.execute("INSERT INTO doc_register (engagement_id,client_id,doc_type,doc_name,doc_category,doc_date,received_from,sent_to,reference_no,status,remarks,created_by_id) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
               (d.get("engagement_id"), d.get("client_id"), d["doc_type"], d["doc_name"], d.get("doc_category"), d["doc_date"], d.get("received_from"), d.get("sent_to"), d.get("reference_no"), d.get("status","Received"), d.get("remarks"), g.user["id"]))
    db.commit(); did = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    log_action(db, g.user["id"], "CREATE_DOC_ENTRY", "DocRegister", did, f"{d['doc_type']}: {d['doc_name']}", request.remote_addr)
    return jsonify(_doc_out(dict_row(db.execute("SELECT * FROM doc_register WHERE id=?", (did,)).fetchone()), db)), 201

@app.route("/api/doc-register/<int:did>", methods=["PUT"])
@login_required
def update_doc_entry(did):
    d = request.get_json(); db = get_db()
    fields, vals = [], []
    for f in ["doc_name","doc_category","doc_date","received_from","sent_to","reference_no","status","remarks","engagement_id","client_id"]:
        if f in d: fields.append(f"{f}=?"); vals.append(d[f])
    if not fields: return jsonify({"detail":"Nothing to update"}), 400
    vals.append(did)
    db.execute(f"UPDATE doc_register SET {','.join(fields)} WHERE id=?", vals); db.commit()
    return jsonify(_doc_out(dict_row(db.execute("SELECT * FROM doc_register WHERE id=?", (did,)).fetchone()), db))

@app.route("/api/doc-register/<int:did>", methods=["DELETE"])
@login_required
def delete_doc_entry(did):
    db = get_db()
    db.execute("DELETE FROM doc_register WHERE id=?", (did,)); db.commit()
    return jsonify({"message":"Deleted"})


# ═══════════════════════════════════════════════════════════════
#  COMPLIANCE / STATUTORY CALENDAR
# ═══════════════════════════════════════════════════════════════
def _cal_out(c):
    return {k: c[k] for k in ["id","title","category","due_date","description","financial_year","engagement_id","is_recurring","status","created_at"]}

@app.route("/api/compliance-calendar/")
@login_required
def list_compliance():
    db = get_db()
    sql = "SELECT * FROM compliance_calendar WHERE 1=1"
    params = []
    if request.args.get("category"):
        sql += " AND category=?"; params.append(request.args["category"])
    if request.args.get("status"):
        sql += " AND status=?"; params.append(request.args["status"])
    if request.args.get("financial_year"):
        sql += " AND financial_year=?"; params.append(request.args["financial_year"])
    sql += " ORDER BY due_date ASC"
    return jsonify([_cal_out(r) for r in dict_rows(db.execute(sql, params).fetchall())])

@app.route("/api/compliance-calendar/", methods=["POST"])
@require_role("Admin","Team Leader")
def create_compliance():
    d = request.get_json(); db = get_db()
    db.execute("INSERT INTO compliance_calendar (title,category,due_date,description,financial_year,engagement_id,is_recurring,status,created_by_id) VALUES (?,?,?,?,?,?,?,?,?)",
               (d["title"], d["category"], d["due_date"], d.get("description"), d.get("financial_year"), d.get("engagement_id"), 1 if d.get("is_recurring") else 0, d.get("status","Upcoming"), g.user["id"]))
    db.commit(); cid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
    log_action(db, g.user["id"], "CREATE_COMPLIANCE", "Compliance", cid, f"{d['category']}: {d['title']}", request.remote_addr)
    return jsonify(_cal_out(dict_row(db.execute("SELECT * FROM compliance_calendar WHERE id=?", (cid,)).fetchone()))), 201

@app.route("/api/compliance-calendar/<int:cid>", methods=["PUT"])
@require_role("Admin","Team Leader")
def update_compliance(cid):
    d = request.get_json(); db = get_db()
    fields, vals = [], []
    for f in ["title","category","due_date","description","financial_year","engagement_id","status","is_recurring"]:
        if f in d: fields.append(f"{f}=?"); vals.append(d[f])
    if not fields: return jsonify({"detail":"Nothing to update"}), 400
    vals.append(cid)
    db.execute(f"UPDATE compliance_calendar SET {','.join(fields)} WHERE id=?", vals); db.commit()
    return jsonify(_cal_out(dict_row(db.execute("SELECT * FROM compliance_calendar WHERE id=?", (cid,)).fetchone())))

@app.route("/api/compliance-calendar/<int:cid>", methods=["DELETE"])
@require_role("Admin","Team Leader")
def delete_compliance(cid):
    db = get_db()
    db.execute("DELETE FROM compliance_calendar WHERE id=?", (cid,)); db.commit()
    return jsonify({"message":"Deleted"})

@app.route("/api/compliance-calendar/seed-fy", methods=["POST"])
@require_role("Admin","Team Leader")
def seed_compliance_fy():
    """Seed standard ITR/GST/TDS/ROC due dates for a financial year."""
    d = request.get_json(); db = get_db()
    fy = d.get("financial_year","")
    if not fy: return jsonify({"detail":"financial_year required"}), 400
    # Parse FY like "2024-25"
    try:
        start_yr = int(fy.split("-")[0])
        end_yr = start_yr + 1
    except Exception:
        return jsonify({"detail":"Invalid FY format. Use YYYY-YY e.g. 2024-25"}), 400
    def yy(y): return str(y)
    deadlines = [
        # GST
        ("GSTR-1 (Quarterly) Q1", "GST", f"{yy(end_yr-1)}-07-31"),
        ("GSTR-1 (Quarterly) Q2", "GST", f"{yy(end_yr-1)}-10-31"),
        ("GSTR-1 (Quarterly) Q3", "GST", f"{yy(end_yr)}-01-31"),
        ("GSTR-1 (Quarterly) Q4", "GST", f"{yy(end_yr)}-04-30"),
        ("GSTR-3B Filing (Monthly)", "GST", f"{yy(end_yr)}-03-20"),
        ("GSTR-9 Annual Return", "GST", f"{yy(end_yr)}-12-31"),
        # TDS
        ("TDS Q1 Return (Form 24Q/26Q)", "TDS", f"{yy(end_yr-1)}-07-31"),
        ("TDS Q2 Return (Form 24Q/26Q)", "TDS", f"{yy(end_yr-1)}-10-31"),
        ("TDS Q3 Return (Form 24Q/26Q)", "TDS", f"{yy(end_yr)}-01-31"),
        ("TDS Q4 Return (Form 24Q/26Q)", "TDS", f"{yy(end_yr)}-05-31"),
        ("TDS Certificate Issue (Form 16)", "TDS", f"{yy(end_yr)}-06-15"),
        # ITR
        ("ITR Filing - Individual/HUF", "ITR", f"{yy(end_yr)}-07-31"),
        ("ITR Filing - Companies / Tax Audit", "ITR", f"{yy(end_yr)}-10-31"),
        ("ITR Filing - Transfer Pricing", "ITR", f"{yy(end_yr)}-11-30"),
        ("Advance Tax Q1 (15%)", "ITR", f"{yy(end_yr-1)}-06-15"),
        ("Advance Tax Q2 (45%)", "ITR", f"{yy(end_yr-1)}-09-15"),
        ("Advance Tax Q3 (75%)", "ITR", f"{yy(end_yr-1)}-12-15"),
        ("Advance Tax Q4 (100%)", "ITR", f"{yy(end_yr)}-03-15"),
        # ROC
        ("ROC Annual Return (MGT-7)", "ROC", f"{yy(end_yr)}-11-29"),
        ("ROC Financial Statements (AOC-4)", "ROC", f"{yy(end_yr)}-10-29"),
        ("AGM Holding Deadline", "ROC", f"{yy(end_yr)}-09-30"),
        ("DIR-3 KYC Filing", "ROC", f"{yy(end_yr)}-09-30"),
    ]
    added = 0
    for title, cat, due in deadlines:
        exists = db.execute("SELECT id FROM compliance_calendar WHERE title=? AND financial_year=?", (title, fy)).fetchone()
        if not exists:
            db.execute("INSERT INTO compliance_calendar (title,category,due_date,financial_year,is_recurring,status,created_by_id) VALUES (?,?,?,?,1,'Upcoming',?)",
                       (title, cat, due, fy, g.user["id"]))
            added += 1
    db.commit()
    return jsonify({"message": f"Seeded {added} compliance dates for FY {fy}"})


# ═══════════════════════════════════════════════════════════════
#  ENHANCED REPORTING & ANALYTICS
# ═══════════════════════════════════════════════════════════════
@app.route("/api/reports/overview")
@login_required
def report_overview():
    db = get_db()
    def cnt(sql, p=()): return db.execute(sql, p).fetchone()[0] or 0
    # Task completion by engagement
    eng_stats = dict_rows(db.execute("""
        SELECT e.id, e.title, c.name as client_name, e.financial_year, e.status as eng_status,
               COUNT(t.id) as total_tasks,
               SUM(CASE WHEN t.status='Completed' OR t.status='Approved' THEN 1 ELSE 0 END) as done_tasks,
               SUM(CASE WHEN t.status='Pending' THEN 1 ELSE 0 END) as pending_tasks,
               SUM(CASE WHEN t.due_date < date('now') AND t.status NOT IN ('Completed','Approved') THEN 1 ELSE 0 END) as overdue_tasks,
               COALESCE(SUM(tl.hours),0) as total_hours
        FROM engagements e
        JOIN clients c ON e.client_id=c.id
        LEFT JOIN tasks t ON t.engagement_id=e.id
        LEFT JOIN time_logs tl ON tl.task_id=t.id
        GROUP BY e.id ORDER BY e.created_at DESC""").fetchall())
    # Staff performance
    staff = dict_rows(db.execute("""
        SELECT u.id, u.full_name, u.role,
               COUNT(DISTINCT t.id) as assigned_tasks,
               SUM(CASE WHEN t.status IN ('Completed','Approved') THEN 1 ELSE 0 END) as completed_tasks,
               COALESCE(SUM(tl.hours),0) as logged_hours,
               SUM(CASE WHEN t.due_date < date('now') AND t.status NOT IN ('Completed','Approved') THEN 1 ELSE 0 END) as overdue_tasks
        FROM users u
        LEFT JOIN tasks t ON t.assignee_id=u.id
        LEFT JOIN time_logs tl ON tl.user_id=u.id
        WHERE u.role IN ('Admin','Team Leader','Member')
        GROUP BY u.id ORDER BY completed_tasks DESC""").fetchall())
    # Query TAT (avg days open→closed)
    tat = db.execute("""SELECT AVG(julianday(responded_date) - julianday(raised_date)) as avg_tat
        FROM queries WHERE responded_date IS NOT NULL""").fetchone()[0]
    # Invoice summary
    inv = dict_row(db.execute("""SELECT COALESCE(SUM(total_amount),0) as billed,
        COALESCE(SUM(CASE WHEN payment_status='Paid' THEN total_amount ELSE 0 END),0) as received,
        COALESCE(SUM(CASE WHEN payment_status='Unpaid' THEN total_amount ELSE 0 END),0) as outstanding
        FROM invoices""").fetchone())
    # Upcoming compliance
    upcoming = dict_rows(db.execute("""SELECT * FROM compliance_calendar
        WHERE status='Upcoming' AND due_date BETWEEN date('now') AND date('now','+30 days')
        ORDER BY due_date LIMIT 10""").fetchall())
    # Tasks due within 7 days
    due_soon = dict_rows(db.execute("""SELECT t.id, t.title, t.due_date, t.status, t.priority,
        u.full_name as assignee_name, e.title as engagement_title
        FROM tasks t
        LEFT JOIN users u ON t.assignee_id=u.id
        LEFT JOIN engagements e ON t.engagement_id=e.id
        WHERE t.due_date BETWEEN date('now') AND date('now','+7 days')
        AND t.status NOT IN ('Completed','Approved')
        ORDER BY t.due_date LIMIT 20""").fetchall())
    return jsonify({
        "engagement_stats": [dict(r) for r in eng_stats],
        "staff_performance": [dict(r) for r in staff],
        "query_avg_tat_days": round(tat, 1) if tat else None,
        "invoice_summary": dict(inv),
        "upcoming_compliance": [_cal_out(r) for r in upcoming],
        "tasks_due_soon": [dict(r) for r in due_soon],
    })


# ═══════════════════════════════════════════════════════════════
#  DUE DATE ALERTS (for frontend polling)
# ═══════════════════════════════════════════════════════════════
@app.route("/api/alerts/")
@login_required
def get_alerts():
    db = get_db(); uid = g.user["id"]
    alerts = []
    # Overdue tasks (assigned to me or all if admin/TL)
    if g.user["role"] in ("Admin","Team Leader"):
        overdue = dict_rows(db.execute("""SELECT t.id, t.title, t.due_date, u.full_name as assignee_name,
            e.title as eng_title FROM tasks t
            LEFT JOIN users u ON t.assignee_id=u.id LEFT JOIN engagements e ON t.engagement_id=e.id
            WHERE t.due_date < date('now') AND t.status NOT IN ('Completed','Approved')
            ORDER BY t.due_date LIMIT 10""").fetchall())
    else:
        overdue = dict_rows(db.execute("""SELECT t.id, t.title, t.due_date, u.full_name as assignee_name,
            e.title as eng_title FROM tasks t
            LEFT JOIN users u ON t.assignee_id=u.id LEFT JOIN engagements e ON t.engagement_id=e.id
            WHERE t.due_date < date('now') AND t.status NOT IN ('Completed','Approved')
            AND (t.assignee_id=? OR t.id IN (SELECT task_id FROM task_assignees WHERE user_id=?))
            ORDER BY t.due_date LIMIT 10""", (uid, uid)).fetchall())
    for t in overdue:
        alerts.append({"type":"overdue_task","severity":"high","message":f"OVERDUE: {t['title']} (due {t['due_date']})","task_id":t["id"],"due_date":t["due_date"]})
    # Tasks due within 3 days
    due_soon = dict_rows(db.execute("""SELECT t.id, t.title, t.due_date FROM tasks t
        WHERE t.due_date BETWEEN date('now') AND date('now','+3 days')
        AND t.status NOT IN ('Completed','Approved')
        AND (t.assignee_id=? OR t.id IN (SELECT task_id FROM task_assignees WHERE user_id=?) OR ? IN ('Admin','Team Leader'))
        ORDER BY t.due_date LIMIT 10""", (uid, uid, g.user["role"])).fetchall())
    for t in due_soon:
        alerts.append({"type":"due_soon","severity":"warn","message":f"Due soon: {t['title']} on {t['due_date']}","task_id":t["id"],"due_date":t["due_date"]})
    # Upcoming compliance (7 days)
    comp = dict_rows(db.execute("""SELECT * FROM compliance_calendar
        WHERE status='Upcoming' AND due_date BETWEEN date('now') AND date('now','+7 days')
        ORDER BY due_date LIMIT 5""").fetchall())
    for c in comp:
        alerts.append({"type":"compliance","severity":"info","message":f"Compliance due: {c['title']} on {c['due_date']}","compliance_id":c["id"],"due_date":c["due_date"]})
    # Open queries older than 7 days
    stale_q = dict_rows(db.execute("""SELECT q.id, q.sr_no, q.query_text, q.raised_date FROM queries q
        WHERE q.status='Open' AND q.raised_date < datetime('now','-7 days') LIMIT 5""").fetchall())
    for q in stale_q:
        alerts.append({"type":"stale_query","severity":"warn","message":f"Query #{q['sr_no']} open >7 days: {q['query_text'][:50]}...","query_id":q["id"]})
    return jsonify(alerts)


# ═══════════════════════════════════════════════════════════════
#  STARTUP
# ═══════════════════════════════════════════════════════════════
@app.route("/api/health")
def health_check():
    return jsonify({"status": "ok", "message": "CA FirmHub is running",
                    "firm_name": FIRM_NAME, "firm_reg_no": FIRM_REG_NO})


@app.route("/api/server/shutdown", methods=["POST"])
@login_required
def server_shutdown():
    """
    Called by the frontend on logout to cleanly shut down the server.
    The start.bat / start.sh launcher will detect this and exit.
    """
    import threading
    log_action(get_db(), g.user["id"], "SERVER_SHUTDOWN", "Server", None,
               "User-initiated shutdown on logout", request.remote_addr)
    print(f"\n  [SERVER] Shutdown requested by {g.user['full_name']}. Stopping...")

    def _stop():
        import time, os, signal
        time.sleep(0.8)  # Let the response reach the browser first
        os.kill(os.getpid(), signal.SIGTERM)

    threading.Thread(target=_stop, daemon=True).start()
    return jsonify({"message": "Server shutting down"})


# ═══════════════════════════════════════════════════════════════
#  TOOL 1: TAX AUDIT (SEC 44AB) ANALYZER
# ═══════════════════════════════════════════════════════════════

@app.route("/api/tools/tax-audit/analyze", methods=["POST"])
@login_required
def tax_audit_analyze():
    """
    Accepts uploaded cash/bank Excel/CSV files + client metadata.
    Parses ledgers, eliminates contra/interbank entries, calculates
    5% cash ratio, and returns the full 44AB analysis as JSON.
    """
    import tempfile, io
    import pandas as pd
    try:
        from rapidfuzz import fuzz
    except ImportError:
        fuzz = None

    software  = request.form.get("software", "Tally")
    miracle   = request.form.get("miracle_system", "Standard Accounting")
    nature    = request.form.get("nature", "Business")
    presumptive = request.form.get("presumptive", "No")
    try:
        turnover = float(request.form.get("turnover", "0").replace(",",""))
    except Exception:
        return jsonify({"detail": "Invalid turnover value"}), 400

    cash_files = request.files.getlist("cash_files")
    bank_files = request.files.getlist("bank_files")
    od_files   = request.files.getlist("od_files")

    if not cash_files and not bank_files and not od_files:
        return jsonify({"detail": "Please upload at least one Cash or Bank ledger file."}), 400

    def clean_numeric(val):
        if val is None: return 0.0
        try:
            import pandas as pd
            if pd.isna(val): return 0.0
        except Exception: pass
        if isinstance(val, (int, float)): return float(val)
        val = str(val).replace(",","").replace("₹","").replace("Rs.","").replace(" ","").strip()
        try: return float(val)
        except: return 0.0

    def check_ignore(row_vals):
        keywords = ["opening balance","closing balance","brought forward","carried forward",
                    "balance b/d","balance c/d","balance b/f","balance c/f","total","grand total"]
        for v in row_vals:
            vs = str(v).lower().strip()
            if any(k in vs for k in keywords): return True
        return False

    def read_file(f, source_type, software, miracle_sys):
        fname = f.filename
        try:
            raw = f.read()
            buf = io.BytesIO(raw)
            if fname.lower().endswith(".csv"):
                df_raw = pd.read_csv(buf, header=None)
            else:
                df_raw = pd.read_excel(buf, header=None)
        except Exception as e:
            return None, str(e)

        try:
            # Find header row
            header_idx = -1
            for i, row in df_raw.head(200).iterrows():
                row_str = " ".join(str(x).lower() for x in row.values)
                if any(k in row_str for k in ["date","dt","particulars","narration","particular"]):
                    header_idx = i; break

            if header_idx != -1:
                df = df_raw.iloc[header_idx:].reset_index(drop=True)
                df.columns = df.iloc[0]
                df = df[1:].reset_index(drop=True)
                df.columns = [str(c).strip().lower() for c in df.columns]
            else:
                df = df_raw.copy()
                df.columns = [f"col_{i}" for i in range(len(df.columns))]

            col_map = {}
            for c in df.columns:
                if not col_map.get("Date") and any(x in c for x in ["date","dt","col_0"]): col_map["Date"] = c
                elif not col_map.get("Particulars") and any(x in c for x in ["particulars","narration","desc","detail","ledger","col_2"]): col_map["Particulars"] = c
                elif not col_map.get("VchType") and any(x in c for x in ["vch","type","voucher","col_3"]): col_map["VchType"] = c
                elif not col_map.get("Debit") and any(x in c for x in ["debit","dr","receipt","deposit","inflow","col_5"]): col_map["Debit"] = c
                elif not col_map.get("Credit") and any(x in c for x in ["credit","cr","payment","withdrawal","outflow","col_6"]): col_map["Credit"] = c

            if "Date" not in col_map or "Particulars" not in col_map:
                return None, f"Could not identify Date/Particulars columns in {fname}"

            df["_date"] = pd.to_datetime(df[col_map["Date"]], dayfirst=True, errors="coerce")
            df["_ignore"] = df.apply(lambda r: check_ignore(r.values), axis=1)
            df["_main"] = df["_date"].notna() | df["_ignore"]
            df["_grp"] = df["_main"].cumsum()
            df = df[df["_grp"] > 0]

            merged = []
            for _, gdf in df.groupby("_grp"):
                mr = gdf.iloc[0]
                d = mr["_date"]
                parts = [str(x).strip() for x in gdf[col_map["Particulars"]].values
                         if pd.notna(x) and str(x).strip() not in ["nan",""]]
                narr = " | ".join(parts)
                vch  = str(mr.get(col_map.get("VchType",""), "")).strip()
                dr   = clean_numeric(mr.get(col_map.get("Debit",""), 0))
                cr   = clean_numeric(mr.get(col_map.get("Credit",""), 0))
                ign  = check_ignore(gdf.values.flatten())
                if dr > 0 or cr > 0:
                    merged.append({"Date":d,"Narration":narr,"Debit":dr,"Credit":cr,
                                   "Source":source_type,"File":fname,
                                   "VchType":vch,"Tag":"IGNORE_ROW" if ign else "NORMAL","Match_ID":""})
            return pd.DataFrame(merged), None
        except Exception as e:
            return None, str(e)

    # Load all files
    all_dfs = []
    failed  = []
    for flist, stype in [(cash_files,"CASH"),(bank_files,"BANK"),(od_files,"BANK_OD")]:
        for f in flist:
            if not f or not f.filename: continue
            df, err = read_file(f, stype, software, miracle)
            if df is not None and not df.empty:
                all_dfs.append(df)
            else:
                failed.append(f.filename + (f" ({err})" if err else ""))

    if not all_dfs:
        return jsonify({"detail": f"No valid data found in any uploaded file. Failed: {', '.join(failed)}"}), 400

    df = pd.concat(all_dfs, ignore_index=True).sort_values("Date").reset_index(drop=True)
    df = df[df["Tag"] != "IGNORE_ROW"].copy().reset_index(drop=True)
    df["ID"] = df.index

    match_counter = [1]
    matched_ids   = set()
    matched_pairs = []

    # Phase 1: Cheque return elimination
    return_kw = ["return","bounce","dishonour","reject","reversal","unpaid"]
    for i, row in df.iterrows():
        if df.at[i,"Tag"] != "NORMAL": continue
        if row["Source"] not in ["BANK","BANK_OD"]: continue
        if i in matched_ids: continue
        nl = str(row["Narration"]).lower()
        vl = str(row.get("VchType","")).lower()
        if not any(k in nl for k in return_kw) and not any(k in vl for k in return_kw): continue
        amt = row["Debit"] if row["Debit"] > 0 else row["Credit"]
        is_dr = row["Debit"] > 0
        mask = ((df["Tag"]=="NORMAL") & (df["File"]==row["File"]) &
                (~df.index.isin(matched_ids)) &
                (abs((df["Date"]-row["Date"]).dt.days) <= 15))
        mask &= (df["Credit"] == amt) if is_dr else (df["Debit"] == amt)
        pot = df[mask]
        if pot.empty: continue
        bi = pot["Date"].sub(row["Date"]).abs().idxmin()
        ms = f"R-{match_counter[0]}"
        df.at[i,"Tag"] = "CHEQUE_RETURN"; df.at[bi,"Tag"] = "CHEQUE_RETURN"
        df.at[i,"Match_ID"] = ms; df.at[bi,"Match_ID"] = ms
        matched_ids |= {i, bi}
        matched_pairs.append({"Match_ID":ms,"Type":"CHEQUE_RETURN","Amount":amt,"Reason":"Cheque Return/Bounce"})
        match_counter[0] += 1

    # Phase 2: Explicit contra voucher elimination
    contra_mask = df["VchType"].str.lower().str.contains("contra|ctra", na=False) & (df["Tag"]=="NORMAL")
    dr_ci = df[contra_mask & (df["Debit"]>0)].index
    cr_ci = df[contra_mask & (df["Credit"]>0)].index
    for di in dr_ci:
        if di in matched_ids: continue
        dr = df.loc[di]; amt = dr["Debit"]
        pot = df.loc[cr_ci]
        pot = pot[(~pot.index.isin(matched_ids)) & (pot["Credit"]==amt) &
                  (abs((pot["Date"]-dr["Date"]).dt.days)<=5)]
        if pot.empty: continue
        ci = pot["Date"].sub(dr["Date"]).abs().idxmin()
        cr_row = df.loc[ci]
        if dr["Source"]=="CASH" and cr_row["Source"] in ["BANK","BANK_OD"]: tag="CONTRA"
        elif dr["Source"] in ["BANK","BANK_OD"] and cr_row["Source"]=="CASH": tag="CONTRA"
        elif dr["Source"] in ["BANK","BANK_OD"] and cr_row["Source"] in ["BANK","BANK_OD"] and dr["File"]!=cr_row["File"]: tag="INTERBANK"
        else: continue
        ms = f"C-{match_counter[0]}"
        df.at[di,"Tag"]=tag; df.at[ci,"Tag"]=tag
        df.at[di,"Match_ID"]=ms; df.at[ci,"Match_ID"]=ms
        matched_ids |= {di,ci}
        matched_pairs.append({"Match_ID":ms,"Type":tag,"Amount":amt,"Reason":"Explicit Contra Voucher"})
        match_counter[0] += 1

    # Phase 3: Single-sided contra
    for idx, row in df.iterrows():
        if df.at[idx,"Tag"] != "NORMAL": continue
        vl = str(row.get("VchType","")).lower()
        if "contra" not in vl and "ctra" not in vl: continue
        tag = "CONTRA" if row["Source"]=="CASH" or "cash" in str(row["Narration"]).lower() else "INTERBANK"
        df.at[idx,"Tag"] = tag; df.at[idx,"Match_ID"] = "Single-Sided"

    # Calculate summary
    full_df = df.copy()
    bank = full_df[full_df["Source"].isin(["BANK","BANK_OD"])]
    cash = full_df[full_df["Source"]=="CASH"]

    gross_bank_rx   = float(bank[bank["Debit"]>0]["Debit"].sum())
    bank_contra_rx  = float(bank[(bank["Tag"]=="CONTRA")&(bank["Debit"]>0)]["Debit"].sum())
    bank_inter_rx   = float(bank[(bank["Tag"]=="INTERBANK")&(bank["Debit"]>0)]["Debit"].sum())
    bank_return_rx  = float(bank[(bank["Tag"]=="CHEQUE_RETURN")&(bank["Debit"]>0)]["Debit"].sum())
    net_bank_rx     = gross_bank_rx - bank_contra_rx - bank_inter_rx - bank_return_rx

    gross_bank_pmt  = float(bank[bank["Credit"]>0]["Credit"].sum())
    bank_contra_pmt = float(bank[(bank["Tag"]=="CONTRA")&(bank["Credit"]>0)]["Credit"].sum())
    bank_inter_pmt  = float(bank[(bank["Tag"]=="INTERBANK")&(bank["Credit"]>0)]["Credit"].sum())
    bank_return_pmt = float(bank[(bank["Tag"]=="CHEQUE_RETURN")&(bank["Credit"]>0)]["Credit"].sum())
    net_bank_pmt    = gross_bank_pmt - bank_contra_pmt - bank_inter_pmt - bank_return_pmt

    gross_cash_rx   = float(cash[cash["Debit"]>0]["Debit"].sum())
    cash_contra_rx  = float(cash[(cash["Tag"]=="CONTRA")&(cash["Debit"]>0)]["Debit"].sum())
    net_cash_rx     = gross_cash_rx - cash_contra_rx

    gross_cash_pmt  = float(cash[cash["Credit"]>0]["Credit"].sum())
    cash_contra_pmt = float(cash[(cash["Tag"]=="CONTRA")&(cash["Credit"]>0)]["Credit"].sum())
    net_cash_pmt    = gross_cash_pmt - cash_contra_pmt

    total_rx  = net_bank_rx + net_cash_rx
    total_pmt = net_bank_pmt + net_cash_pmt
    cash_rx_pct  = round((net_cash_rx / total_rx * 100) if total_rx > 0 else 0.0, 2)
    cash_pmt_pct = round((net_cash_pmt / total_pmt * 100) if total_pmt > 0 else 0.0, 2)

    # Audit decision
    audit_status = "Not Applicable"; section = "N/A"
    reason = "Below threshold limits"
    if nature == "Profession":
        if turnover > 5000000:
            audit_status = "Applicable"; section = "Sec 44AB(b)"
            reason = "Gross receipts from profession > ₹50 Lakhs"
    else:
        if presumptive == "Yes (Opted Out)":
            audit_status = "Applicable"; section = "Sec 44AB(e)"
            reason = "Opted out of presumptive taxation (Sec 44AD(4))"
        elif turnover > 100000000:
            audit_status = "Applicable"; section = "Sec 44AB(a)"
            reason = "Total Turnover > ₹10 Crores"
        elif turnover > 10000000:
            if cash_rx_pct <= 5.0 and cash_pmt_pct <= 5.0:
                audit_status = "Not Applicable"; section = "Proviso to Sec 44AB(a)"
                reason = "Turnover > ₹1 Cr but Cash transactions ≤ 5%"
            else:
                audit_status = "Applicable"; section = "Sec 44AB(a)"
                reason = f"Turnover > ₹1 Cr AND Cash Rx ({cash_rx_pct}%) or Pmt ({cash_pmt_pct}%) > 5%"

    # Ledger-wise working
    ledger_working = []
    for (src, fn), fdf in full_df.groupby(["Source","File"]):
        ledger_working.append({
            "book_type": src, "file": fn,
            "gross_rx": float(fdf[fdf["Debit"]>0]["Debit"].sum()),
            "less_contra_rx": float(fdf[(fdf["Tag"]=="CONTRA")&(fdf["Debit"]>0)]["Debit"].sum()),
            "less_inter_rx":  float(fdf[(fdf["Tag"]=="INTERBANK")&(fdf["Debit"]>0)]["Debit"].sum()),
            "less_return_rx": float(fdf[(fdf["Tag"]=="CHEQUE_RETURN")&(fdf["Debit"]>0)]["Debit"].sum()),
            "net_rx": float(fdf[(fdf["Tag"]=="NORMAL")&(fdf["Debit"]>0)]["Debit"].sum()),
            "gross_pmt": float(fdf[fdf["Credit"]>0]["Credit"].sum()),
            "less_contra_pmt": float(fdf[(fdf["Tag"]=="CONTRA")&(fdf["Credit"]>0)]["Credit"].sum()),
            "less_inter_pmt":  float(fdf[(fdf["Tag"]=="INTERBANK")&(fdf["Credit"]>0)]["Credit"].sum()),
            "less_return_pmt": float(fdf[(fdf["Tag"]=="CHEQUE_RETURN")&(fdf["Credit"]>0)]["Credit"].sum()),
            "net_pmt": float(fdf[(fdf["Tag"]=="NORMAL")&(fdf["Credit"]>0)]["Credit"].sum()),
        })

    # Flags: 40A(3) and 269ST
    normal_df = full_df[full_df["Tag"]=="NORMAL"]
    flags = []
    for _, r in normal_df[normal_df["Source"]=="CASH"].iterrows():
        if r["Credit"] > 10000:
            flags.append({"date": str(r["Date"])[:10], "narration": r["Narration"],
                          "amount": float(r["Credit"]), "warning": "Sec 40A(3) — Cash Payment > ₹10,000"})
        if r["Debit"] >= 200000:
            flags.append({"date": str(r["Date"])[:10], "narration": r["Narration"],
                          "amount": float(r["Debit"]), "warning": "Sec 269ST — Cash Receipt ≥ ₹2,00,000"})

    # Normal entries sample (last 200)
    normal_sample = []
    for _, r in normal_df.tail(200).iterrows():
        normal_sample.append({
            "date": str(r["Date"])[:10], "narration": r["Narration"],
            "debit": float(r["Debit"]), "credit": float(r["Credit"]),
            "source": r["Source"], "file": r["File"], "vch_type": r["VchType"]
        })

    return jsonify({
        "summary": {
            "gross_bank_rx": gross_bank_rx, "bank_contra_rx": bank_contra_rx,
            "bank_inter_rx": bank_inter_rx, "bank_return_rx": bank_return_rx,
            "net_bank_rx": net_bank_rx,
            "gross_bank_pmt": gross_bank_pmt, "bank_contra_pmt": bank_contra_pmt,
            "bank_inter_pmt": bank_inter_pmt, "bank_return_pmt": bank_return_pmt,
            "net_bank_pmt": net_bank_pmt,
            "gross_cash_rx": gross_cash_rx, "cash_contra_rx": cash_contra_rx,
            "net_cash_rx": net_cash_rx,
            "gross_cash_pmt": gross_cash_pmt, "cash_contra_pmt": cash_contra_pmt,
            "net_cash_pmt": net_cash_pmt,
            "total_net_rx": total_rx, "total_net_pmt": total_pmt,
            "cash_rx_pct": cash_rx_pct, "cash_pmt_pct": cash_pmt_pct,
            "audit_status": audit_status, "section": section, "reason": reason,
            "eliminated_count": int(len(df[df["Tag"] != "NORMAL"])),
            "failed_files": failed
        },
        "ledger_working": ledger_working,
        "matched_pairs": matched_pairs,
        "flags": flags,
        "normal_entries": normal_sample
    })


@app.route("/api/tools/tax-audit/export-excel", methods=["POST"])
@login_required
def tax_audit_export_excel():
    """Generate and return a formatted Excel report for the 44AB analysis."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    data = request.get_json(silent=True) or {}
    s    = data.get("summary", {})
    lw   = data.get("ledger_working", [])
    client = data.get("client", {})

    def inr(v):
        try:
            f = float(v)
            s2, *d = f"{abs(f):.2f}".partition(".")
            r = ",".join([s2[x-2:x] for x in range(-3,-len(s2),-2)][::-1]+[s2[-3:]]) if len(s2)>3 else s2
            return f"{'−' if f<0 else ''}₹{r}{''.join(d)}"
        except: return "₹0.00"

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "44AB Working Summary"

    navy  = "003366"; gold = "DAA520"; light = "F0F4F8"; red_c = "C53030"; green_c = "1E7E44"
    hdr_f = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill("solid", fgColor=navy)
    bold_f  = Font(name="Arial", bold=True, size=10)
    reg_f   = Font(name="Arial", size=10)
    ctr     = Alignment(horizontal="center", vertical="center")
    right   = Alignment(horizontal="right", vertical="center")
    left_a  = Alignment(horizontal="left",  vertical="center")
    thin    = Side(style="thin", color="D1DAEA")
    bdr     = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hc(ws, row, col, val, fill=None, font=None, align=None):
        c = ws.cell(row=row, column=col, value=val)
        if fill: c.fill = fill
        if font: c.font = font
        if align: c.alignment = align
        c.border = bdr
        return c

    # Title
    ws.merge_cells("A1:B1")
    t = ws.cell(row=1, column=1, value=f"Tax Audit Applicability Report (Sec 44AB)")
    t.font = Font(name="Arial", bold=True, size=14, color=navy)
    t.alignment = ctr
    ws.row_dimensions[1].height = 28

    # Client info
    rows_info = [
        ("Assessee Name", client.get("name","")),
        ("PAN", client.get("pan","")),
        ("Assessment Year", client.get("ay","")),
        ("Nature", client.get("nature","")),
        ("Turnover / Gross Receipts", inr(client.get("turnover",0))),
    ]
    for i, (k, v) in enumerate(rows_info, 2):
        c1 = ws.cell(row=i, column=1, value=k)
        c1.font = Font(name="Arial", bold=True, size=10); c1.fill = PatternFill("solid", fgColor=light)
        c1.alignment = left_a; c1.border = bdr
        c2 = ws.cell(row=i, column=2, value=v)
        c2.font = reg_f; c2.alignment = left_a; c2.border = bdr

    # Summary table
    r = 8
    ws.merge_cells(f"A{r}:B{r}")
    h = ws.cell(row=r, column=1, value="44AB Working Analysis")
    h.font = hdr_f; h.fill = hdr_fill; h.alignment = ctr
    ws.row_dimensions[r].height = 22; r += 1

    sections = [
        ("=== BANK TRANSACTIONS ===", None),
        ("Total Credits (Gross Receipts)", s.get("gross_bank_rx",0)),
        ("  Less: Cheque Returns (Receipts)", s.get("bank_return_rx",0)),
        ("  Less: Cash Deposited — Contra (Receipts)", s.get("bank_contra_rx",0)),
        ("  Less: Interbank Transfers (Receipts)", s.get("bank_inter_rx",0)),
        ("Bank Receipts for 44AB  (A)", s.get("net_bank_rx",0)),
        ("", None),
        ("Total Debits (Gross Payments)", s.get("gross_bank_pmt",0)),
        ("  Less: Cheque Returns (Payments)", s.get("bank_return_pmt",0)),
        ("  Less: Cash Withdrawn — Contra (Payments)", s.get("bank_contra_pmt",0)),
        ("  Less: Interbank Transfers (Payments)", s.get("bank_inter_pmt",0)),
        ("Bank Payments for 44AB  (B)", s.get("net_bank_pmt",0)),
        ("=== CASH TRANSACTIONS ===", None),
        ("Gross Cash Receipts", s.get("gross_cash_rx",0)),
        ("  Less: Cash Withdrawn from Bank (Contra)", s.get("cash_contra_rx",0)),
        ("Cash Receipts for 44AB  (C)", s.get("net_cash_rx",0)),
        ("", None),
        ("Gross Cash Payments", s.get("gross_cash_pmt",0)),
        ("  Less: Cash Deposited in Bank (Contra)", s.get("cash_contra_pmt",0)),
        ("Cash Payments for 44AB  (D)", s.get("net_cash_pmt",0)),
        ("=== FINAL LIMITS ===", None),
        ("Total Applicable Receipts  (A + C)", s.get("total_net_rx",0)),
        ("Total Applicable Payments  (B + D)", s.get("total_net_pmt",0)),
        (f"Cash Receipt %", f"{s.get('cash_rx_pct',0):.2f} %"),
        (f"Cash Payment %", f"{s.get('cash_pmt_pct',0):.2f} %"),
    ]
    for label, val in sections:
        is_sec  = label.startswith("===")
        is_tot  = "44AB" in label or "Total" in label or "%" in label
        is_blank = label == ""
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = Font(name="Arial", bold=is_sec or is_tot, size=10)
        c1.alignment = left_a; c1.border = bdr
        if is_sec: c1.fill = PatternFill("solid", fgColor=light)
        c2 = ws.cell(row=r, column=2,
                     value="" if val is None or is_blank else (val if isinstance(val, str) else inr(val)))
        c2.font = Font(name="Arial", bold=is_tot, size=10)
        c2.alignment = right; c2.border = bdr
        r += 1

    # Conclusion
    r += 1
    ws.merge_cells(f"A{r}:B{r}")
    ch = ws.cell(row=r, column=1, value="Conclusion")
    ch.font = hdr_f; ch.fill = hdr_fill; ch.alignment = ctr; r += 1
    status_color = green_c if s.get("audit_status") == "Not Applicable" else red_c
    for label, val in [("Audit Status", s.get("audit_status","")),
                        ("Applicable Section", s.get("section","")),
                        ("Reason", s.get("reason",""))]:
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = Font(name="Arial", bold=True, size=10)
        c1.fill = PatternFill("solid", fgColor=light); c1.alignment = left_a; c1.border = bdr
        c2 = ws.cell(row=r, column=2, value=val)
        c2.font = Font(name="Arial", bold=(label=="Audit Status"), size=10,
                       color=status_color if label=="Audit Status" else "000000")
        c2.alignment = left_a; c2.border = bdr; r += 1

    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 22

    # Ledger-wise sheet
    if lw:
        ws2 = wb.create_sheet("Ledger-wise Working")
        hdrs2 = ["Book Type","File / Ledger","Gross Rx","Less Contra Rx","Less Inter Rx",
                 "Less Return Rx","Net Rx (44AB)","Gross Pmt","Less Contra Pmt",
                 "Less Inter Pmt","Less Return Pmt","Net Pmt (44AB)"]
        for ci, h_txt in enumerate(hdrs2, 1):
            c = ws2.cell(row=1, column=ci, value=h_txt)
            c.font = hdr_f; c.fill = hdr_fill; c.alignment = ctr; c.border = bdr
        for ri, rec in enumerate(lw, 2):
            vals = [rec.get("book_type",""), rec.get("file",""),
                    rec.get("gross_rx",0), rec.get("less_contra_rx",0),
                    rec.get("less_inter_rx",0), rec.get("less_return_rx",0), rec.get("net_rx",0),
                    rec.get("gross_pmt",0), rec.get("less_contra_pmt",0),
                    rec.get("less_inter_pmt",0), rec.get("less_return_pmt",0), rec.get("net_pmt",0)]
            for ci, v in enumerate(vals, 1):
                c = ws2.cell(row=ri, column=ci, value=v)
                c.font = reg_f; c.alignment = right if ci > 2 else left_a
                c.border = bdr
                if ci > 2: c.number_format = "#,##0.00"
        for ci in range(1, 13):
            ws2.column_dimensions[get_column_letter(ci)].width = 18 if ci > 2 else 14 if ci == 1 else 30

    import tempfile, os as _os
    tmp = tempfile.NamedTemporaryFile(
        suffix=".xlsx", delete=False,
        dir=_os.path.join(_os.path.dirname(__file__), "static"))
    tmp.close()
    wb.save(tmp.name)
    return send_file(
        tmp.name, as_attachment=True,
        download_name="Tax_Audit_44AB_Report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ═══════════════════════════════════════════════════════════════
#  TOOL 2: PF & ESIC CHALLAN READER
# ═══════════════════════════════════════════════════════════════

@app.route("/api/tools/challan/parse", methods=["POST"])
@login_required
def challan_parse():
    """
    Accepts PF or ESIC challan PDFs. Uses the exact same extraction logic
    as the original PF_ESIC_READER_2.py desktop tool — ported faithfully.
    """
    import re, io
    from datetime import datetime as _dt

    fund_type = request.form.get("fund_type", "PF")
    files = request.files.getlist("files")
    if not files:
        return jsonify({"detail": "No files uploaded"}), 400

    # ── Shared helpers (exact copies from original) ──────────────────────

    MONTHS_S = {"jan":1,"feb":2,"mar":3,"apr":4,"may":5,"jun":6,
                "jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12}
    MONTHS_N = ["","Jan","Feb","Mar","Apr","May","Jun",
                "Jul","Aug","Sep","Oct","Nov","Dec"]

    def _ws(txt):
        return re.sub(r'\s+', ' ', txt.strip()) if txt else ""

    def _amt(s):
        if not s: return 0.0
        s = re.sub(r'[₹Rs\s,]', '', str(s))
        try: return float(s)
        except: return 0.0

    def _find(pattern, text, group=1):
        m = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
        return _ws(m.group(group)) if m else ""

    def _find_all(patterns, text):
        for p in patterns:
            v = _find(p, text)
            if v: return v
        return ""

    def _parse_month_year(raw):
        if not raw: return None
        m = re.search(
            r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*[\s\-/,]*(\d{4})',
            raw, re.IGNORECASE)
        if m:
            mon = MONTHS_S.get(m.group(1).lower()[:3], 0)
            return {"m": mon, "y": int(m.group(2))} if mon else None
        m = re.search(r'(\d{1,2})[\-/](\d{4})', raw)
        if m: return {"m": int(m.group(1)), "y": int(m.group(2))}
        return None

    def _scan_month_year(text):
        m = re.search(
            r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*[\s\-/,]*(\d{4})',
            text, re.IGNORECASE)
        if m:
            mon = MONTHS_S.get(m.group(1).lower()[:3], 0)
            return {"m": mon, "y": int(m.group(2))} if mon else None
        return None

    def _fmt_period(my):
        if not my: return ""
        try: return f"{MONTHS_N[my['m']]}-{my['y']}"
        except: return ""

    def _norm_date(s):
        if not s: return ""
        for fmt in ["%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y",
                    "%m/%d/%Y", "%Y-%m-%d", "%d-%b-%Y", "%d/%b/%Y",
                    "%d-%b-%y", "%d/%b/%y", "%d-%B-%Y", "%d/%B/%Y"]:
            try: return _dt.strptime(s.strip(), fmt).strftime("%d/%m/%Y")
            except: continue
        return s

    def _pf_due_date(my):
        if not my: return "N/A"
        try:
            nm = my["m"] + 1 if my["m"] < 12 else 1
            ny = my["y"] if my["m"] < 12 else my["y"] + 1
            return f"15/{nm:02d}/{ny}"
        except: return "N/A"

    def _esic_due_date(my):
        if not my: return "N/A"
        try:
            nm = my["m"] + 1 if my["m"] < 12 else 1
            ny = my["y"] if my["m"] < 12 else my["y"] + 1
            return f"15/{nm:02d}/{ny}"
        except: return "N/A"

    def _disallowance(due_date, pay_date, emp_share):
        if not due_date or due_date == "N/A" or not pay_date: return 0.0
        try:
            if _dt.strptime(pay_date, "%d/%m/%Y") > _dt.strptime(due_date, "%d/%m/%Y"):
                return round(float(emp_share), 2)
        except: pass
        return 0.0

    def _get_text_from_bytes(raw_bytes):
        """Try pdfplumber first, fall back to all 4 pymupdf strategies."""
        text = ""
        try:
            import pdfplumber, io as _io
            with pdfplumber.open(_io.BytesIO(raw_bytes)) as pdf:
                text = "\n".join(p.extract_text() or "" for p in pdf.pages)
        except Exception: pass

        if len(text.strip()) < 60:
            try:
                import fitz, io as _io
                doc = fitz.open(stream=raw_bytes, filetype="pdf")
                pages_text = []
                for page in doc:
                    best = ""
                    # Strategy 1: plain text
                    try:
                        t = page.get_text("text") or ""
                        if len(t.strip()) > len(best.strip()): best = t
                    except: pass
                    # Strategy 2: word-sorted
                    if len(best.strip()) < 60:
                        try:
                            words = page.get_text("words")
                            if words:
                                words.sort(key=lambda w: (round(w[1]/10)*10, w[0]))
                                lines, cur, cy = [], [], -999
                                for w in words:
                                    y = round(w[1]/10)*10
                                    if abs(y - cy) <= 10: cur.append(w[4])
                                    else:
                                        if cur: lines.append(" ".join(cur))
                                        cur, cy = [w[4]], y
                                if cur: lines.append(" ".join(cur))
                                t = "\n".join(lines)
                                if len(t.strip()) > len(best.strip()): best = t
                        except: pass
                    # Strategy 3: blocks
                    if len(best.strip()) < 60:
                        try:
                            blocks = page.get_text("blocks")
                            t = "\n".join(b[4] for b in blocks if b[6] == 0) if blocks else ""
                            if len(t.strip()) > len(best.strip()): best = t
                        except: pass
                    # Strategy 4: rawdict
                    if len(best.strip()) < 60:
                        try:
                            raw = page.get_text("rawdict")
                            chars = []
                            for block in raw.get("blocks", []):
                                for line in block.get("lines", []):
                                    lt = "".join(s.get("text","") for s in line.get("spans",[]))
                                    if lt.strip(): chars.append(lt)
                            t = "\n".join(chars)
                            if len(t.strip()) > len(best.strip()): best = t
                        except: pass
                    pages_text.append(best)
                doc.close()
                text = "\n".join(pages_text)
            except Exception: pass
        return text

    def _get_page_texts_from_bytes(raw_bytes):
        """Return dict {page_num: text} — used for ESIC multi-page parsing."""
        page_texts = {}
        try:
            import pdfplumber, io as _io
            with pdfplumber.open(_io.BytesIO(raw_bytes)) as pdf:
                for pnum, page in enumerate(pdf.pages, 1):
                    page_texts[pnum] = page.extract_text() or ""
        except Exception: pass

        # Strengthen weak pages with pymupdf
        weak = [p for p, t in page_texts.items() if len(t.strip()) < 60]
        if weak:
            try:
                import fitz, io as _io
                doc = fitz.open(stream=raw_bytes, filetype="pdf")
                for pnum in weak:
                    page = doc[pnum - 1]
                    best = ""
                    for method in ["text", "words", "blocks", "rawdict"]:
                        try:
                            if method == "text":
                                t = page.get_text("text") or ""
                            elif method == "words":
                                words = page.get_text("words")
                                if words:
                                    words.sort(key=lambda w: (round(w[1]/10)*10, w[0]))
                                    lines, cur, cy = [], [], -999
                                    for w in words:
                                        y = round(w[1]/10)*10
                                        if abs(y-cy) <= 10: cur.append(w[4])
                                        else:
                                            if cur: lines.append(" ".join(cur))
                                            cur, cy = [w[4]], y
                                    if cur: lines.append(" ".join(cur))
                                    t = "\n".join(lines)
                                else: t = ""
                            elif method == "blocks":
                                blocks = page.get_text("blocks")
                                t = "\n".join(b[4] for b in blocks if b[6]==0) if blocks else ""
                            elif method == "rawdict":
                                raw = page.get_text("rawdict")
                                chars = []
                                for block in raw.get("blocks",[]):
                                    for line in block.get("lines",[]):
                                        lt = "".join(s.get("text","") for s in line.get("spans",[]))
                                        if lt.strip(): chars.append(lt)
                                t = "\n".join(chars)
                            else: t = ""
                            if len(t.strip()) > len(best.strip()):
                                best = t
                            if len(best.strip()) >= 60: break
                        except: continue
                    if len(best.strip()) > len(page_texts.get(pnum,"").strip()):
                        page_texts[pnum] = best
                doc.close()
            except Exception: pass
        return page_texts

    # ── PF Parsing (exact port of parse_pf_challan) ──────────────────────

    def parse_pf_challan_bytes(raw_bytes, fname):
        text = _get_text_from_bytes(raw_bytes)
        rec = {"filename": fname, "fund_type": "PF", "error": None}

        if len(text.strip()) < 30:
            rec["error"] = ("Could not extract text from this PDF. "
                           "Please download the digital PDF directly from the EPFO portal.")
            return rec

        rec["trrn"] = _find_all([
            r'TRRN\s*(?:No\.?)?\s*[:\-]?\s*(\d{8,20})',
            r'TRRN\s*[:\-]?\s*(\d{8,20})',
        ], text)

        rec["est_code"] = _find_all([
            r'Establishment\s+(?:ID|Code|No\.?)\s*[:\-]?\s*([A-Z]{2,5}[A-Z0-9]{5,16})',
            r'Estt\.?\s+(?:ID|Code)\s*[:\-]?\s*([A-Z0-9]{8,20})',
            r'PF\s+Code\s*[:\-]?\s*([A-Z]{2}[A-Z]{2}\d{7}\d{3})',
        ], text)

        rec["est_name"] = _find_all([
            r'Establishment\s+Name\s*[:\-]?\s*([A-Z][A-Z0-9 &.,()/-]{2,80}?)(?:\n|Wage|TRRN|Total|Account)',
            r'Estt\.?\s+Name\s*[:\-]?\s*([A-Z][A-Z0-9 &.,()/-]{2,80}?)(?:\n)',
            r'Name\s+of\s+Establishment\s*[:\-]?\s*([A-Z][A-Z0-9 &.,()/-]{2,80}?)(?:\n)',
        ], text)

        raw_wm = _find_all([
            r'Wage\s+Month\s*[:\-]?\s*([A-Za-z]+[\s\-/,]+\d{4})',
            r'Wage\s+Month\s*[:\-]?\s*(\d{1,2}[\-/]\d{4})',
            r'For\s+the\s+(?:Month|Period)\s+(?:of\s+)?([A-Za-z]+[\s\-,]+\d{4})',
        ], text)
        rec["wage_month"] = _fmt_period(_parse_month_year(raw_wm) or _scan_month_year(text))

        rec["members"] = _find_all([
            r'Total\s+(?:No\.?\s+of\s+)?Members?\s*[:\-]?\s*(\d+)',
            r'No\.?\s+of\s+Members?\s*[:\-]?\s*(\d+)',
            r'Members?\s*[:\-]\s*(\d+)',
        ], text)

        total_wages_raw = _find_all([
            r'(?:Total\s+)?EPF\s+Wages?\s*[:\-]?\s*([\d,]+(?:\.\d{1,2})?)',
            r'(?:Total\s+)?(?:Gross\s+)?Wages?\s*[:\-]?\s*([\d,]+(?:\.\d{1,2})?)',
        ], text)

        # Account-wise amounts
        acc1 = _find_all([
            r'Account\s*\-?\s*1\s*(?:Amount\s*(?:\(Rs\))?)?\s*[:\-]?\s*([\d,]+(?:\.\d{1,2})?)',
            r'A(?:ccount|/c)\.?\s*(?:No\.?\s*)?1\s*[:\-]?\s*([\d,]+(?:\.\d{1,2})?)',
            r'Acc(?:ount)?\.?\s*1\s*[:\-]?\s*([\d,]+(?:\.\d{1,2})?)',
        ], text)

        emp_share_raw = _find_all([
            r'Employee(?:\'?s?)?\s+(?:Share|Contribution)\s*[:\-]?\s*([\d,]+(?:\.\d{1,2})?)',
            r'EE\s+(?:Share|Contribution)\s*[:\-]?\s*([\d,]+(?:\.\d{1,2})?)',
        ], text)

        total_raw = _find_all([
            r'Total\s+Amount\s*(?:\(Rs\)\s*)?[:\-]?\s*([\d,]+(?:\.\d{1,2})?)',
            r'Grand\s+Total\s*[:\-]?\s*([\d,]+(?:\.\d{1,2})?)',
            r'Total\s*[:\-]\s*([\d,]+(?:\.\d{1,2})?)(?:\s*\n|$)',
        ], text)

        pay_date_raw = _find_all([
            r'(?:Payment\s+Date|Date\s+of\s+Payment|Payment\s+Confirmation\s+Date)\s*[:\-]?\s*(\d{1,2}[/\-][A-Za-z0-9]{2,4}[/\-]\d{2,4})',
            r'(?:Date\s+of\s+Payment|Payment\s+Date|Remittance\s+Date)\s*[:\-]?\s*(\d{1,2}[/\-]\d{1,2}[/\-]\d{4})',
            r'(?:Transaction\s+Date|Paid\s+on|Date\s+of\s+Transaction)\s*[:\-]?\s*(\d{1,2}[/\-]\d{1,2}[/\-]\d{4})',
            r'Date\s*[:\-]?\s*(\d{2}[/\-]\d{2}[/\-]\d{4})',
        ], text)

        wages  = _amt(total_wages_raw)
        acc1_v = _amt(acc1)
        rec["total"] = round(_amt(total_raw), 2)

        if emp_share_raw:
            rec["emp_share"] = round(_amt(emp_share_raw), 2)
        elif wages:
            rec["emp_share"] = round(wages * 0.12, 2)
        elif acc1_v:
            rec["emp_share"] = round(acc1_v * (12 / 15.67), 2)
        else:
            rec["emp_share"] = 0.0

        rec["pay_date"] = _norm_date(pay_date_raw)
        wm_dict = _parse_month_year(raw_wm) or _scan_month_year(text)
        rec["due_date"] = _pf_due_date(wm_dict)
        rec["disallow"] = _disallowance(rec["due_date"], rec["pay_date"], rec["emp_share"])
        return rec

    # ── ESIC Page Parser (exact port of _parse_esic_page) ────────────────

    def _parse_esic_page(t, fname, page_num):
        rec = {"filename": fname, "fund_type": "ESIC", "error": None, "page": page_num}

        ec = _find_all([
            r"Employer(?:'?s?)?\s+Code\s+(?:No\.?|Number|#)\s*[:\-]?\s*(\d[\d\s]{10,20})",
            r"Employer(?:'?s?)?\s+Code\s+No\s*:\s*(\d[\d\s]{10,20})",
            r"Employer\s+Code\s*[:\-\*]?\s*(\d[\d\s]{10,20})",
            r"Code\s+No\.?\s*[:\-]\s*(\d{14,17})",
            r"(\d{17})",
        ], t)
        rec["emp_code"] = ec.replace(" ", "") if ec else ""

        rec["emp_name"] = _find_all([
            r"Employer(?:'?s?)?\s+Name\s*[:\-]\s*([A-Za-z][A-Za-z0-9 &.,()/'/-]{2,80}?)(?:\s*\n|\s{2,}|Challan|Code)",
            r"Name\s+of\s+(?:Employer|Establishment)\s*[:\-]\s*([A-Za-z][A-Za-z0-9 &.,()/'/-]{2,80}?)(?:\s*\n|\s{2,})",
            r"Establishment\s+Name\s*[:\-]\s*([A-Za-z][A-Za-z0-9 &.,()/'/-]{2,80}?)(?:\s*\n|\s{2,}|Challan)",
        ], t)

        rec["challan_no"] = _find_all([
            r"Challan\s+(?:Number|No\.?|#)\s*[:\-]?\s*(\d{8,20})",
            r"Challan\s+ID\s*[:\-]?\s*(\d{8,20})",
        ], t)

        raw_cp = _find_all([
            r"Challan\s+Period\s*[:\-]\s*([\w\-]+)",
            r"(?:Contribution\s+)?Period\s*[:\-]\s*([A-Za-z]{3}[\-/]\d{4})",
            r"(?:Month|Period)\s+of\s+Contribution\s*[:\-]\s*([A-Za-z\-0-9 ]+?)(?:\n|$)",
            r"Wage\s+Month\s*[:\-]\s*([A-Za-z]{3}[\-/\s]\d{4})",
        ], t)

        period_dict = _parse_month_year(raw_cp) or _scan_month_year(t)

        # Fallback: infer from challan created date
        created_raw = _find_all([
            r"Challan\s+Created\s+Date\s+(\d{1,2}[/\-]\d{1,2}[/\-]\d{4})",
            r"Created\s+(?:on\s+)?(\d{1,2}[/\-]\d{1,2}[/\-]\d{4})",
            r"Created\s+Date\s*[:\-]\s*(\d{1,2}[/\-]\d{1,2}[/\-]\d{4})",
        ], t)
        if not period_dict and created_raw:
            try:
                cd_str = created_raw.split()[0]
                cd = _dt.strptime(cd_str, "%d-%m-%Y")
                m = cd.month - 1 if cd.month > 1 else 12
                y = cd.year if cd.month > 1 else cd.year - 1
                period_dict = {"m": m, "y": y}
            except: pass

        rec["period"] = _fmt_period(period_dict)

        rec["pay_date_raw"] = _find_all([
            r"Challan\s+Submitted\s+Date\s+(\d{1,2}[/\-]\d{1,2}[/\-]\d{4})",
            r"Challan\s+Submitted\s+Date\s*[:\-]\s*(\d{1,2}[/\-]\d{1,2}[/\-]\d{4})",
            r"Submitted\s+Date\s*[:\-]\s*(\d{1,2}[/\-]\d{1,2}[/\-]\d{4})",
            r"(?:Date\s+of\s+Payment|Payment\s+Date)\s*[:\-]\s*(\d{1,2}[/\-]\d{1,2}[/\-]\d{4})",
            r"Paid\s+(?:on|On)\s*[:\-]?\s*(\d{1,2}[/\-]\d{1,2}[/\-]\d{4})",
        ], t)

        total_raw = _find_all([
            r"Amount\s+Paid\s*[:\-]\s*([\d,]+(?:\.\d{1,2})?)",
            r"Total\s+Amount\s*[:\-]\s*([\d,]+(?:\.\d{1,2})?)",
            r"Challan\s+Amount\s*[:\-]\s*([\d,]+(?:\.\d{1,2})?)",
            r"Amount\s*[:\-]\s*([\d,]+(?:\.\d{1,2})?)",
        ], t)

        rec["txn_no"] = _find_all([
            r"Transaction\s+(?:Number|No\.?|ID)\s*[:\-]\s*([A-Z0-9]{5,30})",
            r"Txn\.?\s+(?:No\.?|ID|Number)\s*[:\-]\s*([A-Z0-9]{5,30})",
            r"Bank\s+Ref(?:erence)?\s+(?:No\.?|ID)\s*[:\-]\s*([A-Z0-9]{5,30})",
        ], t)

        emp_share_raw = _find_all([
            r"Employee(?:'?s?)?\s+(?:Share|Contribution)\s*[:\-]\s*([\d,]+(?:\.\d{1,2})?)",
            r"EE\s+(?:Share|Contribution)\s*[:\-]\s*([\d,]+(?:\.\d{1,2})?)",
            r"IP\s+(?:Share|Contribution)\s*[:\-]\s*([\d,]+(?:\.\d{1,2})?)",
        ], t)

        wages_raw = _find_all([
            r"Total\s+Wages?\s*[:\-]\s*([\d,]+(?:\.\d{1,2})?)",
            r"Gross\s+Wages?\s*[:\-]\s*([\d,]+(?:\.\d{1,2})?)",
            r"Total\s+Monthly\s+Wages?\s*[:\-]\s*([\d,]+(?:\.\d{1,2})?)",
        ], t)

        rec["ips"] = _find_all([
            r"(?:No\.?\s+of\s+)?(?:IPs?|Insured\s+Persons?)\s*[:\-]\s*(\d+)",
            r"(?:Total\s+)?(?:No\.?\s+of\s+)?Employees?\s*[:\-]\s*(\d+)",
        ], t)

        # Skip pages with no usable data
        if not any([rec["challan_no"], rec["emp_code"], total_raw]):
            return None

        rec["total"] = round(_amt(total_raw), 2)
        rec["emp_share"] = round(_amt(emp_share_raw), 2) if emp_share_raw else 0.0

        if not rec["emp_share"]:
            if rec["total"]:
                rec["emp_share"] = round(rec["total"] * (0.75 / 4.0), 2)
            elif wages_raw:
                rec["emp_share"] = round(_amt(wages_raw) * 0.0075, 2)

        rec["pay_date"] = _norm_date(rec.pop("pay_date_raw", ""))
        rec["due_date"] = _esic_due_date(period_dict)
        rec["disallow"] = _disallowance(rec["due_date"], rec["pay_date"], rec["emp_share"])
        return rec

    # ── ESIC multi-page parser (exact port of parse_esic_challan) ────────

    def parse_esic_challan_bytes(raw_bytes, fname):
        page_texts = _get_page_texts_from_bytes(raw_bytes)
        if not page_texts:
            return [{"filename": fname, "fund_type": "ESIC",
                     "error": "PDF read error — could not open file.", "page": 0}]

        records = []
        for pnum in sorted(page_texts.keys()):
            text = page_texts[pnum]
            if len(text.strip()) < 30: continue
            rec = _parse_esic_page(text, fname, pnum)
            if rec is not None:
                records.append(rec)

        if not records:
            return [{"filename": fname, "fund_type": "ESIC",
                     "error": ("No ESIC challan data could be extracted. "
                               "Please use a digital PDF downloaded from the ESIC portal."),
                     "page": 0}]
        return records

    # ── Main loop ─────────────────────────────────────────────────────────

    results = []
    for f in files:
        if not f or not f.filename: continue
        fname = f.filename
        try:
            raw_bytes = f.read()
            if fund_type == "PF":
                results.append(parse_pf_challan_bytes(raw_bytes, fname))
            else:
                page_recs = parse_esic_challan_bytes(raw_bytes, fname)
                results.extend(page_recs)
        except Exception as e:
            results.append({"filename": fname, "fund_type": fund_type, "error": str(e)})

    return jsonify({"records": results})



@app.route("/api/tools/challan/export-excel", methods=["POST"])
@login_required
def challan_export_excel():
    """
    Generate Form 3CD Clause 20(b) Excel.
    Saves to a temp file (avoids Flask 3.x BytesIO send_file issues).
    Fixes openpyxl MergedCell write corruption.
    """
    import io, os, tempfile, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime as dt

    data      = request.get_json(silent=True) or {}
    pf_recs   = [r for r in data.get("pf_records",  []) if not r.get("error")]
    esic_recs = [r for r in data.get("esic_records", []) if not r.get("error")]
    entity    = data.get("entity", "")

    # ── Colour palette ──
    NAVY   = "1F3864"; MAROON = "7B2D2D"; GOLD   = "C8A951"
    CREAM  = "FFF8E7"; WHITE  = "FFFFFF"; ALTROW = "FDF6EC"
    G_BG   = "D5F5E3"; R_BG   = "FDEDEC"; G_FG   = "1E8449"
    R_FG   = "C0392B"; GREY   = "F0F0F0"; NOTE   = "F8F4EE"

    def _side(c="BFBFBF"):
        return Side(style="thin", color=c)
    def _bdr():
        s = _side(); return Border(left=s, right=s, top=s, bottom=s)
    def _cell(ws, row, col, val, bg=WHITE, fg="000000", bold=False,
              align="center", sz=10, fmt=None, wrap=False):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(name="Arial", bold=bold, color=fg, size=sz)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        c.border    = _bdr()
        if fmt:
            c.number_format = fmt
        return c

    def _fill_merged_row(ws, row, ncols, bg):
        """Fill background for all cells in a merged row (don't write values to slaves)."""
        for col in range(2, ncols + 1):
            cl = ws.cell(row=row, column=col)
            cl.fill   = PatternFill("solid", fgColor=bg)
            cl.border = _bdr()

    def _merge_title(ws, text, ncols, row, bg, fg, sz=11, height=26):
        ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
        c = ws.cell(row=row, column=1, value=text)
        c.font      = Font(name="Arial", bold=True, size=sz, color=fg)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = _bdr()
        ws.row_dimensions[row].height = height
        _fill_merged_row(ws, row, ncols, bg)

    def _sub_row(ws, row, ncols, text, bg=CREAM, height=18):
        ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
        c = ws.cell(row=row, column=1, value=text)
        c.font      = Font(name="Arial", italic=True, size=9, color=NAVY)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = _bdr()
        ws.row_dimensions[row].height = height
        _fill_merged_row(ws, row, ncols, bg)

    def _note_row(ws, row, ncols, text, height=30):
        ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
        c = ws.cell(row=row, column=1, value=text)
        c.font      = Font(name="Arial", italic=True, size=8, color="555555")
        c.fill      = PatternFill("solid", fgColor=NOTE)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border    = _bdr()
        ws.row_dimensions[row].height = height
        _fill_merged_row(ws, row, ncols, NOTE)

    def _hdr_row(ws, row, headers, widths, height=52):
        for ci, (h, w) in enumerate(zip(headers, widths), 1):
            _cell(ws, row, ci, h, bg=MAROON, fg=WHITE, bold=True,
                  align="center", sz=9, wrap=True)
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[row].height = height

    # ── Build Clause 20(b) sheet ──────────────────────────────────────────
    def build_clause20(ws, pf_valid, esic_valid, entity_name):
        NCOLS = 8
        _merge_title(ws,
            "20b.  Details of contributions received from employees for "
            "various funds as referred to in Section 36(1)(va):",
            NCOLS, 1, bg=NAVY, fg=WHITE, height=26)
        _sub_row(ws, 2, NCOLS,
            f"Name of Assessee: {entity_name}   |   "
            f"Prepared on: {dt.now().strftime('%d/%m/%Y')}")
        _note_row(ws, 3, NCOLS,
            "Note:  PF — Employee share = 12% of EPF wages  |  "
            "ESIC — Employee share = 0.75% of wages  |  "
            "Due Date = 15th of month following contribution period  |  "
            "Disallowance u/s 36(1)(va) = Employee share paid after statutory due date")
        _hdr_row(ws, 4,
            ["SN", "Nature of Fund",
             "Sum Received from\nEmployees (₹)",
             "Due Date of\nPayment",
             "Actual Amount\nPaid (₹)",
             "Actual Date of\nPayment",
             "Disallowance (₹)",
             "Challan / Txn\nReference"],
            [5, 36, 22, 16, 20, 22, 20, 24])

        dr = 5
        for rec in pf_valid:
            late = rec.get("disallow", 0) > 0
            dbg  = R_BG if late else G_BG
            dfg  = R_FG if late else G_FG
            bg   = ALTROW if (dr - 5) % 2 == 0 else WHITE
            _cell(ws, dr, 1, dr - 4, bg=bg, bold=True)
            _cell(ws, dr, 2,
                  f"Employees’ Provident Fund (EPF) — {rec.get('wage_month','')}",
                  bg=bg, align="left")
            _cell(ws, dr, 3, rec.get("emp_share", 0), bg=bg, align="right", fmt="#,##0.00")
            _cell(ws, dr, 4, rec.get("due_date",  ""), bg=bg)
            _cell(ws, dr, 5, rec.get("total",      0), bg=bg, align="right", fmt="#,##0.00")
            _cell(ws, dr, 6, rec.get("pay_date",  ""), bg=bg)
            _cell(ws, dr, 7, rec.get("disallow",   0), bg=dbg,
                  align="right", fmt="#,##0.00", fg=dfg, bold=late)
            _cell(ws, dr, 8, f"TRRN: {rec.get('trrn','')}", bg=bg, align="left", sz=8)
            ws.row_dimensions[dr].height = 20
            dr += 1

        for rec in esic_valid:
            late = rec.get("disallow", 0) > 0
            dbg  = R_BG if late else G_BG
            dfg  = R_FG if late else G_FG
            bg   = ALTROW if (dr - 5) % 2 == 0 else WHITE
            _cell(ws, dr, 1, dr - 4, bg=bg, bold=True)
            _cell(ws, dr, 2,
                  f"Employees’ State Insurance (ESIC) — {rec.get('period','')}",
                  bg=bg, align="left")
            _cell(ws, dr, 3, rec.get("emp_share",  0), bg=bg, align="right", fmt="#,##0.00")
            _cell(ws, dr, 4, rec.get("due_date",  ""), bg=bg)
            _cell(ws, dr, 5, rec.get("total",       0), bg=bg, align="right", fmt="#,##0.00")
            _cell(ws, dr, 6, rec.get("pay_date",  ""), bg=bg)
            _cell(ws, dr, 7, rec.get("disallow",   0), bg=dbg,
                  align="right", fmt="#,##0.00", fg=dfg, bold=late)
            _cell(ws, dr, 8,
                  f"Challan: {rec.get('challan_no','')}  Txn: {rec.get('txn_no','')}",
                  bg=bg, align="left", sz=8)
            ws.row_dimensions[dr].height = 20
            dr += 1

        data_end = dr - 1
        # Total row — ONLY write to column A (merged anchor), fill rest without writing values
        ws.merge_cells(f"A{dr}:B{dr}")
        _cell(ws, dr, 1, "TOTAL", bg=NAVY, fg=GOLD, bold=True)
        _fill_merged_row(ws, dr, 2, NAVY)   # fill col B (merged slave) safely
        for ci in range(3, NCOLS + 1):
            if ci in (3, 5, 7):
                ltr = get_column_letter(ci)
                _cell(ws, dr, ci, f"=SUM({ltr}5:{ltr}{data_end})",
                      bg=NAVY, fg=GOLD, bold=True, align="right", fmt="#,##0.00")
            else:
                _cell(ws, dr, ci, "", bg=NAVY)
        ws.row_dimensions[dr].height = 22
        dr += 2

        # Legend
        ws.merge_cells(f"A{dr}:{get_column_letter(NCOLS)}{dr}")
        leg = ws.cell(row=dr, column=1,
                      value=("  ✓ Green = Paid on or before Due Date — "
                             "No Disallowance          "
                             "✗ Red = Paid after Due Date — "
                             "Disallowance u/s 36(1)(va) Applicable"))
        leg.font      = Font(name="Arial", size=8, italic=True, color="333333")
        leg.fill      = PatternFill("solid", fgColor=GREY)
        leg.alignment = Alignment(horizontal="left", vertical="center")
        leg.border    = _bdr()
        ws.row_dimensions[dr].height = 16
        _fill_merged_row(ws, dr, NCOLS, GREY)
        ws.freeze_panes = "A5"

    # ── Build PF Payment Register sheet ──────────────────────────────────
    def build_pf_register(ws, pf_valid, entity_name):
        NCOLS = 10
        _merge_title(ws, f"PF (EPFO) Challan Payment Register — {entity_name}",
                     NCOLS, 1, bg=NAVY, fg=WHITE, height=26)
        _sub_row(ws, 2, NCOLS,
            f"Form 3CD Clause 20(b)  |  Section 36(1)(va)  |  "
            f"Generated: {dt.now().strftime('%d/%m/%Y %H:%M')}")
        _hdr_row(ws, 3,
            ["SN","Wage Month","Establishment Code","Establishment Name",
             "Members","Employee\nContribution (₹)","Total Amount\nPaid (₹)",
             "Due Date","Actual Date\nof Payment","TRRN"],
            [5, 12, 20, 30, 10, 20, 20, 14, 16, 24], height=36)
        for sn, rec in enumerate(pf_valid, 1):
            rn   = sn + 3
            bg   = ALTROW if sn % 2 == 0 else WHITE
            late = rec.get("disallow", 0) > 0
            dbg  = R_BG if late else bg
            dfg  = R_FG if late else "000000"
            _cell(ws, rn, 1,  sn,                           bg=bg, bold=True)
            _cell(ws, rn, 2,  rec.get("wage_month", ""),     bg=bg)
            _cell(ws, rn, 3,  rec.get("est_code",   ""),     bg=bg)
            _cell(ws, rn, 4,  rec.get("est_name",   ""),     bg=bg, align="left")
            _cell(ws, rn, 5,  rec.get("members",    ""),     bg=bg)
            _cell(ws, rn, 6,  rec.get("emp_share",  0),      bg=bg, align="right", fmt="#,##0.00")
            _cell(ws, rn, 7,  rec.get("total",      0),      bg=bg, align="right", fmt="#,##0.00")
            _cell(ws, rn, 8,  rec.get("due_date",   ""),     bg=bg)
            _cell(ws, rn, 9,  rec.get("pay_date",   ""),     bg=dbg, fg=dfg, bold=late)
            _cell(ws, rn, 10, rec.get("trrn",       ""),     bg=bg)
            ws.row_dimensions[rn].height = 20
        de = len(pf_valid) + 3
        ws.merge_cells(f"A{de+1}:E{de+1}")
        _cell(ws, de+1, 1, "TOTAL", bg=NAVY, fg=GOLD, bold=True)
        for ci in range(2, 6): _fill_merged_row(ws, de+1, ci, NAVY) if ci <= 5 else None
        for ci in (6, 7):
            ltr = get_column_letter(ci)
            _cell(ws, de+1, ci, f"=SUM({ltr}4:{ltr}{de})",
                  bg=NAVY, fg=GOLD, bold=True, align="right", fmt="#,##0.00")
        for ci in range(8, NCOLS+1):
            _cell(ws, de+1, ci, "", bg=NAVY)
        ws.row_dimensions[de+1].height = 22
        ws.freeze_panes = "A4"

    # ── Build ESIC Payment Register sheet ────────────────────────────────
    def build_esic_register(ws, esic_valid, entity_name):
        NCOLS = 10
        _merge_title(ws, f"ESIC Challan Payment Register — {entity_name}",
                     NCOLS, 1, bg=NAVY, fg=WHITE, height=26)
        _sub_row(ws, 2, NCOLS,
            f"Form 3CD Clause 20(b)  |  Section 36(1)(va)  |  Emp Rate: 0.75%  |  "
            f"Generated: {dt.now().strftime('%d/%m/%Y %H:%M')}")
        _hdr_row(ws, 3,
            ["SN","Challan Period","Challan Number",
             "Total Amount\nPaid (₹)","Employee Share\n0.75% (₹)",
             "Employer Share\n3.25% (₹)","Due Date",
             "Actual Date\nof Payment","Transaction No.","Employer Code"],
            [5, 12, 18, 18, 18, 18, 14, 16, 18, 22], height=44)
        for sn, rec in enumerate(esic_valid, 1):
            rn    = sn + 3
            bg    = ALTROW if sn % 2 == 0 else WHITE
            late  = rec.get("disallow", 0) > 0
            dbg   = R_BG if late else bg
            dfg   = R_FG if late else "000000"
            total = rec.get("total", 0)
            emp   = round(total * 0.75 / 4.0, 2)
            er    = round(total * 3.25 / 4.0, 2)
            _cell(ws, rn, 1,  sn,                         bg=bg, bold=True)
            _cell(ws, rn, 2,  rec.get("period",     ""),  bg=bg)
            _cell(ws, rn, 3,  rec.get("challan_no", ""),  bg=bg)
            _cell(ws, rn, 4,  total,                      bg=bg, align="right", fmt="#,##0.00")
            _cell(ws, rn, 5,  emp,                        bg=bg, align="right", fmt="#,##0.00")
            _cell(ws, rn, 6,  er,                         bg=bg, align="right", fmt="#,##0.00")
            _cell(ws, rn, 7,  rec.get("due_date",  ""),   bg=bg)
            _cell(ws, rn, 8,  rec.get("pay_date",  ""),   bg=dbg, fg=dfg, bold=late)
            _cell(ws, rn, 9,  rec.get("txn_no",    ""),   bg=bg)
            _cell(ws, rn, 10, rec.get("emp_code",  ""),   bg=bg, sz=8)
            ws.row_dimensions[rn].height = 20
        de = len(esic_valid) + 3
        ws.merge_cells(f"A{de+1}:C{de+1}")
        _cell(ws, de+1, 1, "TOTAL", bg=NAVY, fg=GOLD, bold=True)
        for ci in (4, 5, 6):
            ltr = get_column_letter(ci)
            _cell(ws, de+1, ci, f"=SUM({ltr}4:{ltr}{de})",
                  bg=NAVY, fg=GOLD, bold=True, align="right", fmt="#,##0.00")
        for ci in range(7, NCOLS+1):
            _cell(ws, de+1, ci, "", bg=NAVY)
        ws.row_dimensions[de+1].height = 22
        ws.freeze_panes = "A4"

    # ── Assemble workbook ─────────────────────────────────────────────────
    entity_name = (entity
                   or (pf_recs[0].get("est_name","")   if pf_recs   else "")
                   or (esic_recs[0].get("emp_name","") if esic_recs else ""))

    wb  = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Form 3CD Clause 20(b)"
    build_clause20(ws1, pf_recs, esic_recs, entity_name)

    if pf_recs:
        build_pf_register(wb.create_sheet("PF Payment Register"), pf_recs, entity_name)
    if esic_recs:
        build_esic_register(wb.create_sheet("ESIC Payment Register"), esic_recs, entity_name)

    # ── Save to temp file (Flask 3.x send_file is reliable with real files) ──
    tmp = tempfile.NamedTemporaryFile(
        suffix=".xlsx", delete=False,
        dir=os.path.join(os.path.dirname(__file__), "static")
    )
    tmp.close()
    wb.save(tmp.name)
    return send_file(
        tmp.name,
        as_attachment=True,
        download_name="PF_ESIC_Challan_Clause20_Report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )



if __name__ == "__main__":
    import socket, sys, webbrowser, threading as thr, platform, traceback

    def _fatal(msg, exc=None):
        """Show a clear error message and keep the window open."""
        print("\n" + "!"*60)
        print("  STARTUP ERROR")
        print("!"*60)
        print(f"\n  {msg}\n")
        if exc:
            traceback.print_exc()
        print("\n" + "!"*60)
        print("  Screenshot this window and send to your administrator.")
        print("!"*60 + "\n")
        input("  Press Enter to exit...")
        sys.exit(1)

    try:

        # ── Database initialisation ──────────────────────────────────
        # On the CA firm's machine, the DB may not exist yet (first run).
        # We always initialise — it is safe to call repeatedly (CREATE IF NOT EXISTS).
        with app.app_context():
            init_db()
            db = get_db()
            init_subscriptions_table(db)
            init_v5_tables(db)

            # On first run with a personalised package, register the baked-in subscription
            # into the local DB automatically so verification works offline.
            if FIRM_SUB_ID:
                existing = db.execute(
                    "SELECT id FROM subscription_ids WHERE sub_id=?", (FIRM_SUB_ID,)
                ).fetchone()
                if not existing:
                    db.execute(
                        "INSERT INTO subscription_ids (sub_id, label, firm_name, firm_reg_no, is_active, expires_at, activated) "
                        "VALUES (?,?,?,?,1,?,0)",
                        (FIRM_SUB_ID, f"Package for {FIRM_NAME}", FIRM_NAME, FIRM_REG_NO, FIRM_EXPIRES)
                    )
                    db.commit()
                    print(f"  [INIT] Registered subscription for {FIRM_NAME} ({FIRM_REG_NO})")

            seed(db)
            db.close()
            try:
                from flask import g as _g
                _g.pop('db', None)
            except Exception:
                pass

        # ── Detect LAN IP ────────────────────────────────────────────
        lan_ip = "127.0.0.1"
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.connect(("8.8.8.8", 80))
            lan_ip = s.getsockname()[0]
            s.close()
        except Exception:
            pass

        # ── Windows: add firewall rule ────────────────────────────────
        if platform.system() == "Windows":
            try:
                python_path = sys.executable
                os.system(f'netsh advfirewall firewall delete rule name="CAFirmHub" >nul 2>&1')
                os.system(f'netsh advfirewall firewall add rule name="CAFirmHub" dir=in action=allow protocol=TCP localport={PORT} program="{python_path}" >nul 2>&1')
            except Exception:
                pass

        # ── Check port availability ───────────────────────────────────
        try:
            test_sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            test_sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            test_sock.bind(("0.0.0.0", PORT))
            test_sock.close()
        except OSError:
            print(f"\n  [ERROR] Port {PORT} is already in use!")
            print(f"  Another program is using port {PORT}.")
            print(f"  Either close that program, or set a different port:")
            print(f"    Windows:  set PORT=9000 && python main.py")
            print(f"    Linux:    PORT=9000 python3 main.py")
            sys.exit(1)

        # ── Subscription status display ────────────────────────────────
        firm_display = FIRM_NAME or "(Admin Mode)"
        expiry_display = FIRM_EXPIRES or "Unlimited"

        print(f"""
    ╔═══════════════════════════════════════════════════════════════╗
    ║                   AUDIT MANAGEMENT SYSTEM                     ║
    ║                        v4 Personalised                        ║
    ╠═══════════════════════════════════════════════════════════════╣
    ║                                                               ║
    ║  Firm     : {firm_display:<49}║
    ║  Reg. No. : {FIRM_REG_NO or '—':<49}║
    ║  Sub. Exp : {expiry_display:<49}║
    ║                                                               ║
    ║  Server running at:                                           ║
    ║    This PC:  http://127.0.0.1:{PORT}                          ║
    ║    LAN URL:  http://{lan_ip}:{PORT:<25}                       ║
    ║                                                               ║
    ║  Default Credentials:                                         ║
    ║    Admin:    admin / admin123                                 ║
    ║    Leader:   team.leader / audit123                           ║
    ║    Member:   member1 / audit123                               ║
    ║                                                               ║
    ║  The server stops automatically when you log out.             ║
    ║  To stop manually: Press Ctrl+C                               ║
    ╚═══════════════════════════════════════════════════════════════╝
    """)

        # Auto-open browser
        def open_browser():
            import time
            time.sleep(1.5)
            webbrowser.open(f"http://127.0.0.1:{PORT}")
        thr.Thread(target=open_browser, daemon=True).start()

        # Use waitress (production) or Flask dev server
        try:
            from waitress import serve
            print(f"  [SERVER] Using Waitress (production server)")
            print(f"  [SERVER] Listening on http://0.0.0.0:{PORT}\n")
            serve(app, host=HOST, port=PORT, threads=8)
        except ImportError:
            print(f"  [SERVER] Using Flask development server")
            print(f"  [TIP] For better performance: pip install waitress\n")
            app.run(host=HOST, port=PORT, debug=False, threaded=True)

    except Exception as _e:
        _fatal(f"Unexpected error: {_e}", exc=_e)
    except SystemExit:
        pass

